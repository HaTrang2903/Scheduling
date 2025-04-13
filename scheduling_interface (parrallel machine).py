import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from collections import defaultdict
import matplotlib.cm as cm
import matplotlib.ticker as ticker
import io
import base64
import PIL.Image
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

st.set_page_config(page_title="Job Scheduling Algorithms", layout="wide")

# Title and introduction
st.title("Job Scheduling Algorithms (Parrellel Machine Model)")

# Algorithm selection
with st.sidebar:
    st.header("Select Algorithm")
    algorithm = st.sidebar.selectbox(
        "Choose a scheduling algorithm",
        [
            "Shortest Remaining Processing Time (SRPT)",
            "Shortest Processing Time (SPT)",
            "Longest Processing Time (LPT)",
            "McNaughton's Algorithm"
        ]
    )
    
    st.markdown("---")
    
    # Machine configuration
    st.header("Machine Configuration")
    num_machines = st.number_input("Number of machines", min_value=1, value=1, step=1)
    
    # Only show overlap option for SRPT with multiple machines
    allow_overlap = False
    if algorithm == "Shortest Remaining Processing Time (SRPT)" and num_machines > 1:
        allow_overlap = st.checkbox("Allow job overlap between machines", value=False)

# Input job data section
st.header("Input Job Data")

# Choose input method
input_method = st.radio(
    "Choose input method:",
    ["Upload CSV/Excel", "Manual Input"],
    horizontal=True
)

# Define empty DataFrame for jobs
df = pd.DataFrame(columns=["Job_ID", "Release_Date", "Processing_Time"])

# Upload CSV/Excel option
if input_method == "Upload CSV/Excel":
    uploaded_file = st.file_uploader("Upload job data file", type=["csv", "xlsx", "xls"], 
                                    help="File should have columns: Job_ID, Release_Date, Processing_Time")
    
    if uploaded_file is not None:
        try:
            if uploaded_file.name.endswith('.csv'):
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)
            
            # Validate columns
            required_cols = ["Job_ID", "Release_Date", "Processing_Time"]
            if not all(col in df.columns for col in required_cols):
                st.error(f"File must contain the columns: {', '.join(required_cols)}")
                df = pd.DataFrame(columns=required_cols)
                
        except Exception as e:
            st.error(f"Error reading file: {e}")

# Manual input option
elif input_method == "Manual Input":
    st.write("Enter job details manually:")
    
    # Container for the form
    with st.container():
        num_jobs = st.number_input("Number of jobs", min_value=1, max_value=20, value=3)
        
        # Create a DataFrame for manual input
        manual_data = []
        
        for i in range(int(num_jobs)):
            st.markdown(f"**Job {i+1}**")
            cols = st.columns(3)
            
            job_id = cols[0].text_input(f"Job ID", value=f"J{i+1}", key=f"job_id_{i}")
            release_date = cols[1].number_input(f"Release Date", min_value=0, value=0, key=f"release_{i}")
            processing_time = cols[2].number_input(f"Processing Time", min_value=1, value=i+3, key=f"proc_{i}")
            
            manual_data.append({
                "Job_ID": job_id,
                "Release_Date": release_date,
                "Processing_Time": processing_time
            })
        
        df = pd.DataFrame(manual_data)

# Display the job data
if not df.empty:
    st.subheader("Job Data")
    st.dataframe(df)

# Helper function for Excel download with embedded image
def get_excel_download_link(results_df, fig, algorithm_name):
    """Creates a downloadable Excel file with job results and the Gantt chart"""
    # Create a BytesIO object to store the Excel file
    output = io.BytesIO()
    
    # Create Excel writer
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Add dataframe as a sheet
        results_df.to_excel(writer, sheet_name='Job Results', index=False)
        
        # Add an empty sheet for the Gantt chart
        workbook = writer.book
        gantt_sheet = workbook.create_sheet(title='Gantt Chart')
        
        # Add a title to the Gantt chart sheet
        gantt_sheet['A1'] = f"{algorithm_name} Schedule Gantt Chart"
        
        # Apply styling to the title using proper openpyxl styles
        gantt_sheet['A1'].font = Font(size=14, bold=True)
        gantt_sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Merge cells for the title
        gantt_sheet.merge_cells('A1:G1')
        
        # Save the figure as a PNG in memory
        img_buf = io.BytesIO()
        fig.savefig(img_buf, format='png', dpi=150, bbox_inches='tight')
        img_buf.seek(0)
        
        # Create a PIL Image object from the buffer
        pil_img = PIL.Image.open(img_buf)
        
        # Create an openpyxl image
        xl_img = XLImage(img_buf)
        
        # Resize the image if needed (optional)
        scale_factor = 0.8  # Adjust this for sizing
        xl_img.width = int(pil_img.width * scale_factor)
        xl_img.height = int(pil_img.height * scale_factor)
        
        # Add the image to the sheet
        gantt_sheet.add_image(xl_img, 'A3')
        
        # Auto-adjust columns width for all sheets
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            for column in ws.columns:
                column_letter = get_column_letter(column[0].column)
                if column_letter == 'A':  # Adjust first column which often has headers
                    ws.column_dimensions[column_letter].width = 20
                else:
                    ws.column_dimensions[column_letter].width = 15
    
    # Set buffer position to start
    output.seek(0)
    
    # Convert to base64 for download link
    b64 = base64.b64encode(output.getvalue()).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{algorithm_name}_results.xlsx">📊 Download Results (Excel)</a>'
    
    return href

# Implementation of scheduling algorithms
def srpt_scheduler(jobs, num_machines=1, allow_overlap=False):
    job_dict = {}
    for j in jobs:
        job_dict[j['Job_ID']] = {
            'release': j['release'],
            'processing': j['processing']
        }

    remaining_time = {j['Job_ID']: j['processing'] for j in jobs}
    completed = {j['Job_ID']: False for j in jobs}
    earliest_release = min(j['release'] for j in jobs)
    max_time = earliest_release + sum(j['processing'] for j in jobs)

    schedule = []
    current_time = earliest_release

    while current_time <= max_time:
        available_jobs = [
            job_id for job_id in remaining_time
            if (not completed[job_id])
            and (job_dict[job_id]['release'] <= current_time)
            and (remaining_time[job_id] > 0)
        ]

        if not available_jobs:
            current_time += 1
            continue

        available_jobs.sort(key=lambda j_id: remaining_time[j_id])

        if num_machines == 1:
            job_to_run = available_jobs[0]
            # time, machine, job
            schedule.append((current_time, 1, job_to_run))
            remaining_time[job_to_run] -= 1
            if remaining_time[job_to_run] <= 0:
                completed[job_to_run] = True

        else:
            if not allow_overlap:
                chosen_jobs = available_jobs[:num_machines]
                for m_idx, job_id in enumerate(chosen_jobs, start=1):
                    schedule.append((current_time, m_idx, job_id))
                    remaining_time[job_id] -= 1
                    if remaining_time[job_id] <= 0:
                        completed[job_id] = True
            else:
                machine_used = 0
                idx_jobs = 0
                while machine_used < num_machines and idx_jobs < len(available_jobs):
                    job_id = available_jobs[idx_jobs]
                    needed_for_this_job = remaining_time[job_id]
                    can_alloc = min(needed_for_this_job, num_machines - machine_used)
                    for m in range(can_alloc):
                        schedule.append((current_time, machine_used + 1 + m, job_id))
                    
                    remaining_time[job_id] -= can_alloc
                    machine_used += can_alloc
                    
                    if remaining_time[job_id] <= 0:
                        completed[job_id] = True
                    
                    idx_jobs += 1

        current_time += 1

        if all(completed.values()):
            break

    completion_time = {}
    for job_id in job_dict:
        job_entries = [(t, m, j) for (t, m, j) in schedule if j == job_id]
        if job_entries:
            completion_time[job_id] = max([t for (t, m, j) in job_entries]) + 1
        else:
            completion_time[job_id] = None

    # Calculate makespan (max completion time across all jobs)
    makespan = max(completion_time.values()) if completion_time else 0

    return {
        'schedule': schedule,
        'completion_time': completion_time,
        'makespan': makespan
    }

def spt_scheduler(jobs, num_machines):
    # Sort jobs by processing time (shortest first)
    sorted_jobs = sorted(jobs, key=lambda x: x['processing'])
    
    # Initialize schedule
    schedule = []
    
    # Track machine end times
    machine_end_times = [0] * num_machines
    
    # Assign jobs to machines
    for job in sorted_jobs:
        job_id = job['Job_ID']
        release_time = job['release']
        processing_time = job['processing']
        
        # Find the machine that will finish earliest
        earliest_machine = min(range(num_machines), key=lambda m: machine_end_times[m])
        
        # Calculate start time (max of machine availability and job release time)
        start_time = max(machine_end_times[earliest_machine], release_time)
        
        # Create scheduling entries for this job
        for t in range(int(start_time), int(start_time + processing_time)):
            schedule.append((t, earliest_machine + 1, job_id))
        
        # Update machine end time
        machine_end_times[earliest_machine] = start_time + processing_time
    
    # Calculate completion times for each job
    completion_times = {}
    for job_id in set(j['Job_ID'] for j in sorted_jobs):
        relevant_entries = [entry for entry in schedule if entry[2] == job_id]
        if relevant_entries:
            completion_times[job_id] = max([entry[0] for entry in relevant_entries]) + 1
        else:
            completion_times[job_id] = None
    
    # Calculate makespan (max completion time across all machines)
    makespan = max(machine_end_times) if machine_end_times else 0
    
    return {
        'schedule': schedule,
        'completion_time': completion_times,
        'makespan': makespan
    }

def lpt_scheduler(jobs, num_machines):
    # Sort jobs by processing time (longest first)
    sorted_jobs = sorted(jobs, key=lambda x: x['processing'], reverse=True)
    
    # Initialize schedule
    schedule = []
    
    # Track machine end times
    machine_end_times = [0] * num_machines
    
    # Assign jobs to machines
    for job in sorted_jobs:
        job_id = job['Job_ID']
        release_time = job['release']
        processing_time = job['processing']
        
        # Find the machine that will finish earliest
        earliest_machine = min(range(num_machines), key=lambda m: machine_end_times[m])
        
        # Calculate start time (max of machine availability and job release time)
        start_time = max(machine_end_times[earliest_machine], release_time)
        
        # Create scheduling entries for this job
        for t in range(int(start_time), int(start_time + processing_time)):
            schedule.append((t, earliest_machine + 1, job_id))
        
        # Update machine end time
        machine_end_times[earliest_machine] = start_time + processing_time
    
    # Calculate completion times for each job
    completion_times = {}
    for job_id in set(j['Job_ID'] for j in sorted_jobs):
        relevant_entries = [entry for entry in schedule if entry[2] == job_id]
        if relevant_entries:
            completion_times[job_id] = max([entry[0] for entry in relevant_entries]) + 1
        else:
            completion_times[job_id] = None
    
    # Calculate makespan (max completion time across all machines)
    makespan = max(machine_end_times) if machine_end_times else 0
    
    return {
        'schedule': schedule,
        'completion_time': completion_times,
        'makespan': makespan
    }

def mcnaughton_scheduler(jobs, num_machines):
    # Calculate the optimal makespan
    total_processing_time = sum(job['processing'] for job in jobs)
    longest_job = max(job['processing'] for job in jobs)
    
    # Optimal makespan is the maximum of:
    # 1. The longest job processing time
    # 2. The total processing time divided by the number of machines
    optimal_makespan = max(longest_job, total_processing_time / num_machines)
    
    # Sort jobs by processing time (not required but helps visualization)
    sorted_jobs = sorted(jobs, key=lambda x: x['processing'], reverse=True)
    
    # Initialize schedule
    schedule = []
    
    # Keep track of current machine and time
    current_machine = 1
    current_time = 0
    
    # Assign jobs to machines
    for job in sorted_jobs:
        job_id = job['Job_ID']
        release_time = job['release']  # Note: McNaughton's algorithm assumes all jobs are available at time 0
        processing_time = job['processing']
        remaining_time = processing_time
        
        while remaining_time > 0:
            # Time available on current machine before reaching optimal_makespan
            available_time = optimal_makespan - current_time
            
            # If job fits completely in current machine
            if remaining_time <= available_time:
                # Schedule the job on current machine
                for t in range(int(current_time), int(current_time + remaining_time)):
                    schedule.append((t, current_machine, job_id))
                
                # Update current time
                current_time += remaining_time
                remaining_time = 0
                
                # If we've reached the optimal makespan, move to next machine
                if current_time >= optimal_makespan:
                    current_machine += 1
                    current_time = 0
            
            # If job doesn't fit completely (needs to be split)
            else:
                # Schedule part of the job on current machine
                for t in range(int(current_time), int(optimal_makespan)):
                    schedule.append((t, current_machine, job_id))
                
                # Update remaining time
                remaining_time -= (optimal_makespan - current_time)
                
                # Move to next machine
                current_machine += 1
                current_time = 0
    
    # Calculate completion times for each job
    completion_times = {}
    for job_id in set(j['Job_ID'] for j in sorted_jobs):
        relevant_entries = [entry for entry in schedule if entry[2] == job_id]
        if relevant_entries:
            completion_times[job_id] = max([entry[0] for entry in relevant_entries]) + 1
        else:
            completion_times[job_id] = None
    
    # Calculate makespan (should be close to optimal_makespan)
    machine_completion_times = [0] * num_machines
    for machine_idx in range(num_machines):
        machine_jobs = [entry for entry in schedule if entry[1] == machine_idx + 1]
        if machine_jobs:
            machine_completion_times[machine_idx] = max(t for t, m, j in machine_jobs) + 1
        else:
            machine_completion_times[machine_idx] = 0
    
    makespan = max(machine_completion_times) if machine_completion_times else 0
    
    return {
        'schedule': schedule,
        'completion_time': completion_times,
        'makespan': makespan,
        'optimal_makespan': optimal_makespan
    }

def build_gantt_chart(schedule, jobs_data, title="Gantt Chart"):
    """Creates a matplotlib figure with the Gantt chart visualization"""
    machine_segments = defaultdict(list)
    sched_sorted = sorted(schedule, key=lambda x: (x[1], x[0]))

    # Create segments for visualization
    for (t, mach, job) in sched_sorted:
        if not machine_segments[mach]:
            machine_segments[mach].append([job, t, t + 1])
        else:
            last = machine_segments[mach][-1]
            if last[0] == job and last[2] == t:
                last[2] = t + 1
            else:
                machine_segments[mach].append([job, t, t + 1])

    # Extract release times from jobs_data
    release_times = {job['Job_ID']: job['release'] for job in jobs_data}

    # Create figure and axis with wider width
    fig, ax = plt.subplots(figsize=(14, 5))

    # Use a better color palette
    all_jobs = sorted(set(j['Job_ID'] for j in jobs_data))
    colors = cm.tab20(np.linspace(0, 1, len(all_jobs)))  # Use tab20 for more colors
    color_map = {job_id: colors[i] for i, job_id in enumerate(all_jobs)}

    # Get all machine numbers and sort them
    all_machines = sorted(machine_segments.keys())

    # Reverse the machine order mapping so Machine 1 is at the top
    machine_positions = {mach: len(all_machines) - idx - 1 for idx, mach in enumerate(all_machines)}

    y_ticks = []
    y_labels = []

    # Find max time for axis setting
    max_time = 0
    for mach in machine_segments:
        for segment in machine_segments[mach]:
            max_time = max(max_time, segment[2])

    # Plot job segments with reversed y-axis (Machine 1 at top)
    for mach in all_machines:
        y_pos = machine_positions[mach]
        for segment in machine_segments[mach]:
            job_id, start_t, end_t = segment
            ax.barh(y_pos, end_t - start_t, left=start_t,
                    color=color_map[job_id], edgecolor='black', linewidth=1)

            # Add job label if segment is wide enough
            if end_t - start_t > 0.5:
                ax.text((start_t + end_t) / 2, y_pos, f"{job_id}",
                        va='center', ha='center', color='white',
                        fontsize=10, fontweight='bold')

        y_ticks.append(y_pos)
        y_labels.append(f"Machine {mach}")

    # Add blue vertical lines for release dates (but only at the bottom of the chart)
    for job_id, release_time in release_times.items():
        if release_time > 0:  # Only show non-zero release times
            ax.axvline(x=release_time, color='blue', linestyle=':', linewidth=1, alpha=0.7)

    # Set axis properties
    ax.set_yticks(y_ticks)
    ax.set_yticklabels(y_labels)

    # Fixed tick interval of 2
    tick_interval = 2

    # Set x-axis with ticks every 2 units
    ax.set_xlim(0, max_time)
    ax.set_xticks(np.arange(0, max_time + 1, tick_interval))

    # Make x-axis tick labels smaller to avoid overlap
    ax.tick_params(axis='x', labelsize=8)

    # Add minor ticks and grid for better readability
    ax.xaxis.set_minor_locator(ticker.MultipleLocator(1))  # Minor ticks every 1 unit
    ax.grid(axis='x', which='major', linestyle='-', alpha=0.3)
    ax.grid(axis='x', which='minor', linestyle=':', alpha=0.15)

    # Set labels and title
    ax.set_xlabel("Time", fontweight='bold')
    ax.set_ylabel("Machine", fontweight='bold')
    ax.set_title(title, fontsize=14, fontweight='bold')

    # Remove top and right spines for cleaner look
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    plt.tight_layout()
    
    return fig

# Run scheduling and display results
if not df.empty and st.button("Run Algorithm"):
    # Create a placeholder for results
    result_container = st.container()
    figure_container = st.container()
    download_container = st.container()
    
    # Convert DataFrame to job data format
    jobs_data = []
    for _, row in df.iterrows():
        jobs_data.append({
            'Job_ID': row['Job_ID'],
            'release': row['Release_Date'],
            'processing': row['Processing_Time']
        })
    
    # Run the selected algorithm
    if algorithm == "Shortest Remaining Processing Time (SRPT)":
        result = srpt_scheduler(jobs_data, num_machines=num_machines, allow_overlap=allow_overlap)
        if num_machines == 1:
            chart_title = "Single Machine SRPT"
        else:
            overlap_text = "Overlap Allowed" if allow_overlap else "No Overlap"
            chart_title = f"Multi Machine SRPT ({num_machines} Machines, {overlap_text})"
    
    elif algorithm == "Shortest Processing Time (SPT)":
        result = spt_scheduler(jobs_data, num_machines=num_machines)
        chart_title = f"{'Single' if num_machines == 1 else 'Multi'} Machine SPT ({num_machines} {'Machine' if num_machines == 1 else 'Machines'})"
    
    elif algorithm == "Longest Processing Time (LPT)":
        result = lpt_scheduler(jobs_data, num_machines=num_machines)
        chart_title = f"{'Single' if num_machines == 1 else 'Multi'} Machine LPT ({num_machines} {'Machine' if num_machines == 1 else 'Machines'})"
    
    elif algorithm == "McNaughton's Algorithm":
        result = mcnaughton_scheduler(jobs_data, num_machines=num_machines)
        if num_machines == 1:
            chart_title = "Single Machine McNaughton (equivalent to SPT)"
        else:
            chart_title = f"McNaughton's Algorithm ({num_machines} Machines)"
    
    # Create results DataFrame with just completion times
    results_data = []
    for job in jobs_data:
        job_id = job['Job_ID']
        processing_time = job['processing']
        release_time = job['release']
        completion_time = result['completion_time'].get(job_id, None)
        
        results_data.append({
            'Job_ID': job_id,
            'Processing_Time': processing_time,
            'Release_Date': release_time,
            'Completion_Time': completion_time
        })
    
    results_df = pd.DataFrame(results_data)
    
    # Generate Gantt chart
    fig = build_gantt_chart(result['schedule'], jobs_data, title=chart_title)
    
    # Display results
    with result_container:
        st.subheader(f"Results: {algorithm}")
        
        # Display makespan
        st.metric("Makespan", result['makespan'])
        
        # Display McNaughton specific info if applicable
        if algorithm == "McNaughton's Algorithm" and 'optimal_makespan' in result:
            st.info(f"Theoretical Optimal Makespan: {result['optimal_makespan']:.2f}")
        
        # Display job completion times
        st.subheader("Job Completion Times:")
        st.dataframe(results_df)
    
    # Display the Gantt chart
    with figure_container:
        st.pyplot(fig)
    
    # Add Excel download link
    with download_container:
        st.subheader("Download Results")
        excel_link = get_excel_download_link(results_df, fig, algorithm)
        st.markdown(excel_link, unsafe_allow_html=True)