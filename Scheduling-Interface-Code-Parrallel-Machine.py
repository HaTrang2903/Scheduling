import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from collections import defaultdict
import matplotlib.cm as cm
import matplotlib.ticker as ticker
import io
import base64
import PIL.Image
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

st.set_page_config(page_title="Job Scheduling Algorithms", layout="wide")

# Title and introduction
st.title("Job Scheduling Algorithms (Parallel Machine Model)")

# Algorithm selection
with st.sidebar:
    st.header("Select Algorithm")
    algorithm = st.sidebar.selectbox(
        "Choose a scheduling algorithm",
        [
            "Shortest Remaining Processing Time (SRPT)",
            "Shortest Processing Time (SPT)",
            "Longest Processing Time (LPT)",
            "McNaughton's Algorithm"
        ]
    )

    st.markdown("---")

    # Algorithm Description
    st.header("Algorithm Description")
    
    # Descriptions for each algorithm
    if algorithm == "Shortest Remaining Processing Time (SRPT)":
        st.info("""
        Prioritizes jobs with the shortest remaining processing time. Preemptive algorithm that interrupts execution when a job with shorter remaining time arrives. Minimizes average completion time but may lead to starvation of longer jobs. Optimal for minimizing average flow time on a single machine.
        """)
    elif algorithm == "Shortest Processing Time (SPT)":
        st.info("""
        Arranges jobs in order of processing time, from shortest to longest. Non-preemptive algorithm that minimizes average completion time and average waiting time. Simple to implement but can lead to starvation of longer jobs. Good for systems where quick average response time is important.
        """)
    elif algorithm == "Longest Processing Time (LPT)":
        st.info("""
        Schedules jobs in decreasing order of processing time (longest first). Provides better load balancing across multiple machines and reduces makespan. Not optimal for average completion time but helps prevent processor starvation. Good for batch processing environments where completion of all jobs matters more than individual response times.
        """)
    elif algorithm == "McNaughton's Algorithm":
        st.info("""
        An optimal algorithm for scheduling parallel identical machines to minimize makespan. Creates a schedule with makespan equal to max(longest job, total processing time / machines). May split jobs across machines with no preemption cost. Guarantees optimal makespan for parallel machine scheduling.
        """)

    st.markdown("---")

    # Machine configuration
    st.header("Machine Configuration")
    num_machines = st.number_input("Number of machines", min_value=1, value=1, step=1)

    # Only show overlap option for SRPT with multiple machines
    allow_overlap = False
    if algorithm == "Shortest Remaining Processing Time (SRPT)" and num_machines > 1:
        allow_overlap = st.checkbox("Allow job overlap between machines", value=False)
    
    st.markdown("---")
    
    # Visualization options
    st.header("Visualization Options")
    show_release_lines = st.checkbox("Show release date lines", value=True)

# Input job data section
st.header("Input Job Data")

# Initialize session state for dataframe if it doesn't exist
if 'df' not in st.session_state:
    st.session_state.df = None
    st.session_state.file_uploaded = False

# Choose input method
input_method = st.radio(
    "Choose input method:",
    ["Upload CSV/Excel", "Manual Input"],
    horizontal=True
)

# Upload CSV/Excel option
if input_method == "Upload CSV/Excel":
    uploaded_file = st.file_uploader("Upload job data file", type=["csv", "xlsx", "xls"],
                                    help="File should have columns: Job_ID, Release_Date, Processing_Time")

    # Only load the file if it's newly uploaded
    if uploaded_file is not None and not st.session_state.file_uploaded:
        try:
            if uploaded_file.name.endswith('.csv'):
                st.session_state.df = pd.read_csv(uploaded_file)
            else:
                st.session_state.df = pd.read_excel(uploaded_file)

            # Validate columns
            required_cols = ["Job_ID", "Release_Date", "Processing_Time"]
            if not all(col in st.session_state.df.columns for col in required_cols):
                st.error(f"File must contain the columns: {', '.join(required_cols)}")
                st.session_state.df = pd.DataFrame(columns=required_cols)
            else:
                st.session_state.file_uploaded = True

        except Exception as e:
            st.error(f"Error reading file: {e}")
            st.session_state.df = None

    # Check if we have data to work with from a previous upload
    if st.session_state.df is not None and st.session_state.file_uploaded:
        st.write("Uploaded data:")
        st.dataframe(st.session_state.df)

        # Add data editing options
        st.subheader("Edit Uploaded Data")
        edit_option = st.selectbox("Choose an action:", ["No changes needed", "Edit existing rows", "Add new rows", "Delete rows"])

        if edit_option == "Edit existing rows":
            st.write("Select a row to edit:")
            for i, row in st.session_state.df.iterrows():
                col1, col2 = st.columns([1, 3])
                with col1:
                    edit_this_row = st.checkbox(f"Edit row {i+1}", key=f"edit_{i}")
                with col2:
                    st.write(row.to_dict())

                if edit_this_row:
                    st.write(f"Editing row {i+1}:")
                    edited_row = {}
                    cols = st.columns(3)  # 3 columns for parallel machine data

                    # Job ID column (can be string or number)
                    with cols[0]:
                        edited_row["Job_ID"] = st.text_input(
                            "Job ID",
                            value=str(row["Job_ID"]),
                            key=f"edit_{i}_job_id"
                        )

                    # Release Date column (numeric)
                    with cols[1]:
                        edited_row["Release_Date"] = st.number_input(
                            "Release Date",
                            value=float(row["Release_Date"]),
                            min_value=0.0,
                            step=1.0,
                            key=f"edit_{i}_release_date"
                        )

                    # Processing Time column (numeric)
                    with cols[2]:
                        edited_row["Processing_Time"] = st.number_input(
                            "Processing Time",
                            value=float(row["Processing_Time"]),
                            min_value=1.0,
                            step=1.0,
                            key=f"edit_{i}_processing_time"
                        )

                    if st.button(f"Update Row {i+1}", key=f"update_{i}"):
                        for col_name, val in edited_row.items():
                            st.session_state.df.at[i, col_name] = val
                        st.success(f"Row {i+1} updated!")

        elif edit_option == "Add new rows":
            st.write("Add a new row:")
            new_row = {}
            cols = st.columns(3)

            with cols[0]:
                # Suggest a unique Job ID format if possible
                if not st.session_state.df.empty and all(isinstance(id, str) and id.startswith('J') for id in st.session_state.df['Job_ID']):
                    # If all job IDs are in Jx format, suggest next number
                    try:
                        numbers = [int(id.replace('J', '')) for id in st.session_state.df['Job_ID']]
                        next_num = max(numbers) + 1
                        default_job_id = f"J{next_num}"
                    except:
                        default_job_id = f"J{len(st.session_state.df) + 1}"
                else:
                    default_job_id = f"J{len(st.session_state.df) + 1}"

                new_row["Job_ID"] = st.text_input(
                    "Job ID",
                    value=default_job_id,
                    key="new_job_id"
                )

            with cols[1]:
                new_row["Release_Date"] = st.number_input(
                    "Release Date",
                    value=0.0,
                    min_value=0.0,
                    step=1.0,
                    key="new_release_date"
                )

            with cols[2]:
                new_row["Processing_Time"] = st.number_input(
                    "Processing Time",
                    value=5.0,
                    min_value=1.0,
                    step=1.0,
                    key="new_processing_time"
                )

            if st.button("Add Row", key="add_row_button"):
                # Append the new row to the session state dataframe
                st.session_state.df = pd.concat([st.session_state.df, pd.DataFrame([new_row])], ignore_index=True)
                st.success("New row added!")

        elif edit_option == "Delete rows":
            st.write("Select rows to delete:")
            rows_to_delete = []
            for i, row in st.session_state.df.iterrows():
                col1, col2 = st.columns([1, 3])
                with col1:
                    delete_this_row = st.checkbox(f"Delete row {i+1}", key=f"delete_{i}")
                    if delete_this_row:
                        rows_to_delete.append(i)
                with col2:
                    st.write(row.to_dict())

            if rows_to_delete and st.button("Delete Selected Rows", key="delete_rows_button"):
                st.session_state.df = st.session_state.df.drop(rows_to_delete).reset_index(drop=True)
                st.success(f"{len(rows_to_delete)} row(s) deleted!")

        # Display the updated data
        st.subheader("Current Data")
        st.dataframe(st.session_state.df)

        # Add a button to clear the data if needed
        if st.button("Clear uploaded data and start over", key="clear_data_button"):
            st.session_state.df = None
            st.session_state.file_uploaded = False
            st.rerun()

        # Set df to use the session state data for algorithms
        df = st.session_state.df.copy()

# Manual input option
elif input_method == "Manual Input":
    st.write("Enter job details manually:")

    # If we already have manually entered data, use it
    if 'df' in st.session_state and st.session_state.df is not None and not st.session_state.file_uploaded:
        num_jobs = st.number_input("Number of jobs", min_value=1, max_value=20, value=len(st.session_state.df))

        # If jobs were added or removed, update the dataframe
        if num_jobs != len(st.session_state.df):
            # If increasing, add new rows
            if num_jobs > len(st.session_state.df):
                for i in range(len(st.session_state.df), num_jobs):
                    new_row = {
                        "Job_ID": f"J{i+1}",
                        "Release_Date": 0,
                        "Processing_Time": i+3  # Varied processing times
                    }
                    st.session_state.df = pd.concat([st.session_state.df, pd.DataFrame([new_row])], ignore_index=True)
            # If decreasing, remove rows
            else:
                st.session_state.df = st.session_state.df.iloc[:num_jobs].reset_index(drop=True)

        # Manual editing for each job
        manual_data = []

        for i in range(int(num_jobs)):
            st.markdown(f"**Job {i+1}**")
            cols = st.columns(3)

            current_job = st.session_state.df.iloc[i] if i < len(st.session_state.df) else None

            job_id = cols[0].text_input(
                "Job ID",
                value=current_job["Job_ID"] if current_job is not None else f"J{i+1}",
                key=f"job_id_{i}"
            )

            release_date = cols[1].number_input(
                "Release Date",
                min_value=0.0,
                value=float(current_job["Release_Date"]) if current_job is not None else 0.0,
                key=f"release_{i}"
            )

            processing_time = cols[2].number_input(
                "Processing Time",
                min_value=1.0,
                value=float(current_job["Processing_Time"]) if current_job is not None else float(i+3),
                key=f"proc_{i}"
            )

            manual_data.append({
                "Job_ID": job_id,
                "Release_Date": release_date,
                "Processing_Time": processing_time
            })

        # Update the session state dataframe with the manual data
        st.session_state.df = pd.DataFrame(manual_data)

    else:
        # Start fresh with manual input
        num_jobs = st.number_input("Number of jobs", min_value=1, max_value=20, value=3)

        # Container for the form
        with st.container():
            manual_data = []

            for i in range(int(num_jobs)):
                st.markdown(f"**Job {i+1}**")
                cols = st.columns(3)

                job_id = cols[0].text_input(f"Job ID", value=f"J{i+1}", key=f"job_id_{i}")
                release_date = cols[1].number_input(f"Release Date", min_value=0, value=0, key=f"release_{i}")
                processing_time = cols[2].number_input(f"Processing Time", min_value=1, value=i+3, key=f"proc_{i}")

                manual_data.append({
                    "Job_ID": job_id,
                    "Release_Date": release_date,
                    "Processing_Time": processing_time
                })

            # Create dataframe and store in session state
            st.session_state.df = pd.DataFrame(manual_data)
            st.session_state.file_uploaded = False  # This is manual data, not uploaded

    # Display the current data
    st.subheader("Job Data")
    st.dataframe(st.session_state.df)

    # Set df for algorithms
    df = st.session_state.df.copy()

# Final check to ensure df is properly defined before algorithm execution
if 'df' not in locals() or df is None:
    if 'df' in st.session_state and st.session_state.df is not None:
        df = st.session_state.df.copy()
    else:
        # Create an empty DataFrame with the required columns
        df = pd.DataFrame(columns=["Job_ID", "Release_Date", "Processing_Time"])

# Display run button (but don't show the data again)
if not df.empty:
    # No need to repeat the data display
    run_algorithm = st.button("Run Algorithm", key="run_algo_button")
else:
    st.warning("Please input job data before running the algorithm.")
    run_algorithm = False

# Helper function for Excel download with embedded image
def get_excel_download_link(results_df, fig, algorithm_name):
    """Creates a downloadable Excel file with job results and the Gantt chart"""
    # Create a BytesIO object to store the Excel file
    output = io.BytesIO()

    # Create Excel writer
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Add dataframe as a sheet
        results_df.to_excel(writer, sheet_name='Job Results', index=False)

        # Add an empty sheet for the Gantt chart
        workbook = writer.book
        gantt_sheet = workbook.create_sheet(title='Gantt Chart')

        # Add a title to the Gantt chart sheet
        gantt_sheet['A1'] = f"{algorithm_name} Schedule Gantt Chart"

        # Apply styling to the title using proper openpyxl styles
        gantt_sheet['A1'].font = Font(size=14, bold=True)
        gantt_sheet['A1'].alignment = Alignment(horizontal='center')

        # Merge cells for the title
        gantt_sheet.merge_cells('A1:G1')

        # Save the figure as a PNG in memory
        img_buf = io.BytesIO()
        fig.savefig(img_buf, format='png', dpi=150, bbox_inches='tight')
        img_buf.seek(0)

        # Create a PIL Image object from the buffer
        pil_img = PIL.Image.open(img_buf)

        # Create an openpyxl image
        xl_img = XLImage(img_buf)

        # Resize the image if needed (optional)
        scale_factor = 0.8  # Adjust this for sizing
        xl_img.width = int(pil_img.width * scale_factor)
        xl_img.height = int(pil_img.height * scale_factor)

        # Add the image to the sheet
        gantt_sheet.add_image(xl_img, 'A3')

        # Auto-adjust columns width for all sheets
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            for column in ws.columns:
                column_letter = get_column_letter(column[0].column)
                if column_letter == 'A':  # Adjust first column which often has headers
                    ws.column_dimensions[column_letter].width = 20
                else:
                    ws.column_dimensions[column_letter].width = 15

    # Set buffer position to start
    output.seek(0)

    # Convert to base64 for download link
    b64 = base64.b64encode(output.getvalue()).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{algorithm_name}_results.xlsx">📊 Download Results (Excel)</a>'

    return href

# Implementation of scheduling algorithms
def srpt_scheduler(jobs, num_machines=1, allow_overlap=False):
    job_dict = {}
    for j in jobs:
        job_dict[j['Job_ID']] = {
            'release': j['release'],
            'processing': j['processing']
        }

    remaining_time = {j['Job_ID']: j['processing'] for j in jobs}
    completed = {j['Job_ID']: False for j in jobs}
    earliest_release = min(j['release'] for j in jobs)
    max_time = earliest_release + sum(j['processing'] for j in jobs)

    schedule = []
    current_time = earliest_release

    # Track which job is currently running on each machine
    current_job_on_machine = {m: None for m in range(1, num_machines + 1)}

    while current_time <= max_time:
        available_jobs = [
            job_id for job_id in remaining_time
            if (not completed[job_id])
            and (job_dict[job_id]['release'] <= current_time)
            and (remaining_time[job_id] > 0)
        ]

        if not available_jobs:
            current_time += 1
            continue

        # Primary sort by remaining time, but for equal times, prefer continuing jobs
        if num_machines > 1:
            # Create a custom sort key that prioritizes:
            # 1. Shortest remaining time
            # 2. Jobs already running on a machine
            available_jobs.sort(key=lambda j_id: (
                remaining_time[j_id],
                0 if j_id in current_job_on_machine.values() else 1,
                j_id  # As final tiebreaker, sort alphabetically
            ))
        else:
            # For single machine, just sort by remaining time
            available_jobs.sort(key=lambda j_id: remaining_time[j_id])

        if num_machines == 1:
            job_to_run = available_jobs[0]
            schedule.append((current_time, 1, job_to_run))
            current_job_on_machine[1] = job_to_run
            remaining_time[job_to_run] -= 1
            if remaining_time[job_to_run] <= 0:
                completed[job_to_run] = True
                current_job_on_machine[1] = None

        else:
            if not allow_overlap:
                # Try to assign jobs to machines they were already running on first
                machine_assignment = {}
                assigned_jobs = set()

                # First pass: try to assign jobs to their current machines
                for m in range(1, num_machines + 1):
                    current_job = current_job_on_machine[m]
                    if current_job in available_jobs and current_job not in assigned_jobs:
                        machine_assignment[m] = current_job
                        assigned_jobs.add(current_job)

                # Second pass: assign remaining jobs to empty machines
                for job_id in available_jobs:
                    if job_id not in assigned_jobs:
                        for m in range(1, num_machines + 1):
                            if m not in machine_assignment:
                                machine_assignment[m] = job_id
                                assigned_jobs.add(job_id)
                                break

                # Schedule the assigned jobs
                for m, job_id in machine_assignment.items():
                    schedule.append((current_time, m, job_id))
                    current_job_on_machine[m] = job_id
                    remaining_time[job_id] -= 1
                    if remaining_time[job_id] <= 0:
                        completed[job_id] = True
                        current_job_on_machine[m] = None
            else:
                # Your existing logic for overlapping jobs...
                # (with similar modifications to track current_job_on_machine)
                machine_used = 0
                idx_jobs = 0
                while machine_used < num_machines and idx_jobs < len(available_jobs):
                    job_id = available_jobs[idx_jobs]
                    needed_for_this_job = remaining_time[job_id]
                    can_alloc = min(needed_for_this_job, num_machines - machine_used)
                    for m in range(int(can_alloc)):
                        machine_num = machine_used + 1 + m
                        schedule.append((current_time, machine_num, job_id))
                        current_job_on_machine[machine_num] = job_id

                    remaining_time[job_id] -= can_alloc
                    machine_used += can_alloc

                    if remaining_time[job_id] <= 0:
                        completed[job_id] = True
                        # Clear job from all machines it was running on
                        for m in range(1, num_machines + 1):
                            if current_job_on_machine[m] == job_id:
                                current_job_on_machine[m] = None

                    idx_jobs += 1

        current_time += 1

        if all(completed.values()):
            break

    # Rest of your existing code...
    completion_time = {}
    for job_id in job_dict:
        job_entries = [(t, m, j) for (t, m, j) in schedule if j == job_id]
        if job_entries:
            completion_time[job_id] = max([t for (t, m, j) in job_entries]) + 1
        else:
            completion_time[job_id] = None

    # Calculate makespan (max completion time across all jobs)
    makespan = max(completion_time.values()) if completion_time else 0

    return {
        'schedule': schedule,
        'completion_time': completion_time,
        'makespan': makespan
    }

def spt_scheduler(jobs, num_machines):
    # Sort jobs by processing time (shortest first)
    sorted_jobs = sorted(jobs, key=lambda x: x['processing'])

    # Initialize schedule
    schedule = []

    # Track machine end times
    machine_end_times = [0] * num_machines

    # Assign jobs to machines
    for job in sorted_jobs:
        job_id = job['Job_ID']
        release_time = job['release']
        processing_time = job['processing']

        # Find the machine that will finish earliest
        earliest_machine = min(range(num_machines), key=lambda m: machine_end_times[m])

        # Calculate start time (max of machine availability and job release time)
        start_time = max(machine_end_times[earliest_machine], release_time)

        # Create scheduling entries for this job
        for t in range(int(start_time), int(start_time + processing_time)):
            schedule.append((t, earliest_machine + 1, job_id))

        # Update machine end time
        machine_end_times[earliest_machine] = start_time + processing_time

    # Calculate completion times for each job
    completion_times = {}
    for job_id in set(j['Job_ID'] for j in sorted_jobs):
        relevant_entries = [entry for entry in schedule if entry[2] == job_id]
        if relevant_entries:
            completion_times[job_id] = max([entry[0] for entry in relevant_entries]) + 1
        else:
            completion_times[job_id] = None

    # Calculate makespan (max completion time across all machines)
    makespan = max(machine_end_times) if machine_end_times else 0

    return {
        'schedule': schedule,
        'completion_time': completion_times,
        'makespan': makespan
    }

def lpt_scheduler(jobs, num_machines):
    # Sort jobs by processing time (longest first)
    sorted_jobs = sorted(jobs, key=lambda x: x['processing'], reverse=True)

    # Initialize schedule
    schedule = []

    # Track machine end times
    machine_end_times = [0] * num_machines

    # Assign jobs to machines
    for job in sorted_jobs:
        job_id = job['Job_ID']
        release_time = job['release']
        processing_time = job['processing']

        # Find the machine that will finish earliest
        earliest_machine = min(range(num_machines), key=lambda m: machine_end_times[m])

        # Calculate start time (max of machine availability and job release time)
        start_time = max(machine_end_times[earliest_machine], release_time)

        # Create scheduling entries for this job
        for t in range(int(start_time), int(start_time + processing_time)):
            schedule.append((t, earliest_machine + 1, job_id))

        # Update machine end time
        machine_end_times[earliest_machine] = start_time + processing_time

    # Calculate completion times for each job
    completion_times = {}
    for job_id in set(j['Job_ID'] for j in sorted_jobs):
        relevant_entries = [entry for entry in schedule if entry[2] == job_id]
        if relevant_entries:
            completion_times[job_id] = max([entry[0] for entry in relevant_entries]) + 1
        else:
            completion_times[job_id] = None

    # Calculate makespan (max completion time across all machines)
    makespan = max(machine_end_times) if machine_end_times else 0

    return {
        'schedule': schedule,
        'completion_time': completion_times,
        'makespan': makespan
    }

def mcnaughton_scheduler(jobs, num_machines):
    # Calculate the optimal makespan
    total_processing_time = sum(job['processing'] for job in jobs)
    longest_job = max(job['processing'] for job in jobs)

    # Optimal makespan is the maximum of:
    # 1. The longest job processing time
    # 2. The total processing time divided by the number of machines
    optimal_makespan = max(longest_job, total_processing_time / num_machines)

    # Sort jobs by processing time (not required but helps visualization)
    sorted_jobs = sorted(jobs, key=lambda x: x['processing'], reverse=True)

    # Initialize schedule
    schedule = []

    # Keep track of current machine and time
    current_machine = 1
    current_time = 0

    # Assign jobs to machines
    for job in sorted_jobs:
        job_id = job['Job_ID']
        release_time = job['release']  # Note: McNaughton's algorithm assumes all jobs are available at time 0
        processing_time = job['processing']
        remaining_time = processing_time

        while remaining_time > 0:
            # Time available on current machine before reaching optimal_makespan
            available_time = optimal_makespan - current_time

            # If job fits completely in current machine
            if remaining_time <= available_time:
                # Schedule the job on current machine
                for t in range(int(current_time), int(current_time + remaining_time)):
                    schedule.append((t, current_machine, job_id))

                # Update current time
                current_time += remaining_time
                remaining_time = 0

                # If we've reached the optimal makespan, move to next machine
                if current_time >= optimal_makespan:
                    current_machine += 1
                    current_time = 0

            # If job doesn't fit completely (needs to be split)
            else:
                # Schedule part of the job on current machine
                for t in range(int(current_time), int(optimal_makespan)):
                    schedule.append((t, current_machine, job_id))

                # Update remaining time
                remaining_time -= (optimal_makespan - current_time)

                # Move to next machine
                current_machine += 1
                current_time = 0

    # Calculate completion times for each job
    completion_times = {}
    for job_id in set(j['Job_ID'] for j in sorted_jobs):
        relevant_entries = [entry for entry in schedule if entry[2] == job_id]
        if relevant_entries:
            completion_times[job_id] = max([entry[0] for entry in relevant_entries]) + 1
        else:
            completion_times[job_id] = None

    # Calculate makespan (should be close to optimal_makespan)
    machine_completion_times = [0] * num_machines
    for machine_idx in range(num_machines):
        machine_jobs = [entry for entry in schedule if entry[1] == machine_idx + 1]
        if machine_jobs:
            machine_completion_times[machine_idx] = max(t for t, m, j in machine_jobs) + 1
        else:
            machine_completion_times[machine_idx] = 0

    makespan = max(machine_completion_times) if machine_completion_times else 0

    return {
        'schedule': schedule,
        'completion_time': completion_times,
        'makespan': makespan,
        'optimal_makespan': optimal_makespan
    }

def build_gantt_chart(schedule, jobs_data, title="Gantt Chart", show_release_lines=True):
    """Creates a matplotlib figure with the Gantt chart visualization"""
    machine_segments = defaultdict(list)
    sched_sorted = sorted(schedule, key=lambda x: (x[1], x[0]))

    # Create segments for visualization
    for (t, mach, job) in sched_sorted:
        if not machine_segments[mach]:
            machine_segments[mach].append([job, t, t + 1])
        else:
            last = machine_segments[mach][-1]
            if last[0] == job and last[2] == t:
                last[2] = t + 1
            else:
                machine_segments[mach].append([job, t, t + 1])

    # Extract release times from jobs_data
    release_times = {job['Job_ID']: job['release'] for job in jobs_data}

    # Create figure and axis with wider width
    fig, ax = plt.subplots(figsize=(14, 5))

    # Use a better color palette
    all_jobs = sorted(set(j['Job_ID'] for j in jobs_data))
    colors = cm.tab20(np.linspace(0, 1, len(all_jobs)))  # Use tab20 for more colors
    color_map = {job_id: colors[i] for i, job_id in enumerate(all_jobs)}

    # Get all machine numbers and sort them
    all_machines = sorted(machine_segments.keys())

    # Reverse the machine order mapping so Machine 1 is at the top
    machine_positions = {mach: len(all_machines) - idx - 1 for idx, mach in enumerate(all_machines)}

    y_ticks = []
    y_labels = []
    # Find max time for axis setting
    max_time = 0
    for mach in machine_segments:
        for segment in machine_segments[mach]:
            max_time = max(max_time, segment[2])

    # Plot job segments with reversed y-axis (Machine 1 at top)
    for mach in all_machines:
        y_pos = machine_positions[mach]
        for segment in machine_segments[mach]:
            job_id, start_t, end_t = segment
            ax.barh(y_pos, end_t - start_t, left=start_t,
                    color=color_map[job_id], edgecolor='black', linewidth=1)

            # Add job label if segment is wide enough
            if end_t - start_t > 0.5:
                ax.text((start_t + end_t) / 2, y_pos, f"{job_id}",
                        va='center', ha='center', color='white',
                        fontsize=10, fontweight='bold')

        y_ticks.append(y_pos)
        y_labels.append(f"Machine {mach}")

    # Add blue vertical lines for release dates - only if toggle is enabled
    if show_release_lines:
        for job_id, release_time in release_times.items():
            if release_time > 0:  # Only show non-zero release times
                ax.axvline(x=release_time, color='blue', linestyle=':', linewidth=1, alpha=0.7)

    # Set axis properties
    ax.set_yticks(y_ticks)
    ax.set_yticklabels(y_labels)

    # Fixed tick interval of 2
    tick_interval = 2

    # Set x-axis with ticks every 2 units
    ax.set_xlim(0, max_time)
    ax.set_xticks(np.arange(0, max_time + 1, tick_interval))

    # Make x-axis tick labels smaller to avoid overlap
    ax.tick_params(axis='x', labelsize=8)

    # Add minor ticks and grid for better readability
    ax.xaxis.set_minor_locator(ticker.MultipleLocator(1))  # Minor ticks every 1 unit
    ax.grid(axis='x', which='major', linestyle='-', alpha=0.3)
    ax.grid(axis='x', which='minor', linestyle=':', alpha=0.15)

    # Set labels and title
    ax.set_xlabel("Time", fontweight='bold')
    ax.set_ylabel("Machine", fontweight='bold')
    ax.set_title(title, fontsize=14, fontweight='bold')

    # Remove top and right spines for cleaner look
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    plt.tight_layout()

    return fig

# Run scheduling and display results
if not df.empty and run_algorithm:
    # Create a placeholder for results
    result_container = st.container()
    figure_container = st.container()
    download_container = st.container()

    # Convert DataFrame to job data format
    jobs_data = []
    for _, row in df.iterrows():
        jobs_data.append({
            'Job_ID': row['Job_ID'],
            'release': row['Release_Date'],
            'processing': row['Processing_Time']
        })

    # Run the selected algorithm
    if algorithm == "Shortest Remaining Processing Time (SRPT)":
        result = srpt_scheduler(jobs_data, num_machines=num_machines, allow_overlap=allow_overlap)
        if num_machines == 1:
            chart_title = "Single Machine SRPT"
        else:
            overlap_text = "Overlap Allowed" if allow_overlap else "No Overlap"
            chart_title = f"Multi Machine SRPT ({num_machines} Machines, {overlap_text})"

    elif algorithm == "Shortest Processing Time (SPT)":
        result = spt_scheduler(jobs_data, num_machines=num_machines)
        chart_title = f"{'Single' if num_machines == 1 else 'Multi'} Machine SPT ({num_machines} {'Machine' if num_machines == 1 else 'Machines'})"

    elif algorithm == "Longest Processing Time (LPT)":
        result = lpt_scheduler(jobs_data, num_machines=num_machines)
        chart_title = f"{'Single' if num_machines == 1 else 'Multi'} Machine LPT ({num_machines} {'Machine' if num_machines == 1 else 'Machines'})"

    elif algorithm == "McNaughton's Algorithm":
        result = mcnaughton_scheduler(jobs_data, num_machines=num_machines)
        if num_machines == 1:
            chart_title = "Single Machine McNaughton (equivalent to SPT)"
        else:
            chart_title = f"McNaughton's Algorithm ({num_machines} Machines)"

    # Create results DataFrame with just completion times
    results_data = []
    for job in jobs_data:
        job_id = job['Job_ID']
        processing_time = job['processing']
        release_time = job['release']
        completion_time = result['completion_time'].get(job_id, None)

        results_data.append({
            'Job_ID': job_id,
            'Processing_Time': processing_time,
            'Release_Date': release_time,
            'Completion_Time': completion_time
        })

    results_df = pd.DataFrame(results_data)

    # Generate Gantt chart - pass the show_release_lines parameter
    fig = build_gantt_chart(result['schedule'], jobs_data, title=chart_title, show_release_lines=show_release_lines)

    # Display results
    with result_container:
        st.subheader(f"Results: {algorithm}")

        # Display makespan
        st.metric("Makespan", result['makespan'])

        # Display McNaughton specific info if applicable
        if algorithm == "McNaughton's Algorithm" and 'optimal_makespan' in result:
            st.info(f"Theoretical Optimal Makespan: {result['optimal_makespan']:.2f}")

        # Display job completion times
        st.subheader("Job Completion Times:")
        st.dataframe(results_df)

    # Display the Gantt chart
    with figure_container:
        st.pyplot(fig)

    # Add Excel download link
    with download_container:
        st.subheader("Download Results")
        excel_link = get_excel_download_link(results_df, fig, algorithm)
        st.markdown(excel_link, unsafe_allow_html=True)