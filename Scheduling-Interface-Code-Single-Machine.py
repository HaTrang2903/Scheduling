import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import matplotlib.patches as patches
import io
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
import random
import time
import heapq
import networkx as nx
from matplotlib.lines import Line2D
import base64
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import PIL.Image

# Set page title
st.title("Job Scheduling Algorithms (Single Machine Model)")

# Sidebar for algorithm selection
st.sidebar.header("Select Algorithm")
algorithm = st.sidebar.selectbox(
    "Choose a scheduling algorithm",
    [
        "First Come First Serve (FCFS)",
        "Shortest Processing Time (SPT)",
        "Longest Processing Time (LPT)",
        "Earliest Due Date (EDD)",
        "Earliest Release Date (ERD)",
        "Weighted Shortest Processing Time (WSPT)",
        "Random Sequencing (RAND)",
        "Moore's Rule (Minimize Late Jobs)",
        "Shortest Remaining Processing Time (SRPT)",
        "Branch and Bound (Minimize Maximum Lateness)"
    ]
)

# Algorithm descriptions dictionary
algorithm_descriptions = {
    "First Come First Serve (FCFS)": "Processes jobs in the order they arrive. Simple to implement and fair, but not optimized for performance metrics. Best for systems where arrival order is important.",
    
    "Shortest Processing Time (SPT)": "Prioritizes jobs with the shortest processing times first. Minimizes average flow time and average waiting time. Optimal for reducing mean flow time when all jobs are available at the same time.",
    
    "Longest Processing Time (LPT)": "Schedules jobs with longest processing times first. Can be useful for load balancing multiple machines, though on a single machine it typically increases flow time.",
    
    "Earliest Due Date (EDD)": "Processes jobs in order of their due dates. Minimizes maximum lateness. Best when meeting deadlines is the primary concern.",
    
    "Earliest Release Date (ERD)": "Schedules jobs based on when they become available. Good for scenarios where jobs arrive at different times and must be processed as soon as possible.",
    
    "Weighted Shortest Processing Time (WSPT)": "Orders jobs by ascending processing time/weight ratio. Minimizes weighted flow time and weighted completion time. Use when jobs have different importance or priority levels.",
    
    "Random Sequencing (RAND)": "Schedules jobs in a random order. Serves as a benchmark for comparing other algorithms, with no specific optimization objective.",
    
    "Moore's Rule (Minimize Late Jobs)": "Minimizes the number of late jobs. Greedily schedules jobs by EDD and removes the longest job when a tardy job is encountered. Best when the goal is to complete as many jobs on time as possible.",
    
    "Shortest Remaining Processing Time (SRPT)": "Preemptive version of SPT that always processes the job with shortest remaining time. Minimizes mean flow time when preemption is allowed.",
    
    "Branch and Bound (Minimize Maximum Lateness)": "Exact algorithm that finds the optimal schedule to minimize maximum lateness. Uses a tree search with pruning to find the optimal solution efficiently."
}

# Display the description of the selected algorithm
st.sidebar.markdown("---")
st.sidebar.subheader("Algorithm Description")
st.sidebar.info(algorithm_descriptions[algorithm])
st.sidebar.markdown("---")

# Display the release date lines
st.sidebar.markdown("---")
st.sidebar.header("Visualization Options")
show_release_lines = st.sidebar.checkbox("Show release date lines", value=True)

# Data input section
st.header("Input Job Data")
upload_option = st.radio("Choose input method:", ["Upload CSV/Excel", "Manual Input"])

# Initialize session state for dataframe if it doesn't exist
if 'df' not in st.session_state:
    st.session_state.df = None
    st.session_state.file_uploaded = False

# Initialize df variable to None by default
df = None

if upload_option == "Upload CSV/Excel":
    uploaded_file = st.file_uploader("Upload job data file", type=["csv", "xlsx"])

    # Only load the file if it's newly uploaded
    if uploaded_file is not None and not st.session_state.file_uploaded:
        if uploaded_file.name.endswith('.csv'):
            st.session_state.df = pd.read_csv(uploaded_file)
        else:
            st.session_state.df = pd.read_excel(uploaded_file)
        st.session_state.file_uploaded = True

    # Check if we have data to work with
    if st.session_state.df is not None:
        st.write("Uploaded data:")
        st.dataframe(st.session_state.df)

        # Add data editing options
        st.subheader("Edit Uploaded Data")
        edit_option = st.selectbox("Choose an action:", ["No changes needed", "Edit existing rows", "Add new rows", "Delete rows"])

        if edit_option == "Edit existing rows":
            st.write("Select a row to edit:")
            for i, row in st.session_state.df.iterrows():
                col1, col2 = st.columns([1, 3])
                with col1:
                    edit_this_row = st.checkbox(f"Edit row {i+1}", key=f"edit_{i}")
                with col2:
                    st.write(row.to_dict())

                if edit_this_row:
                    st.write(f"Editing row {i+1}:")
                    edited_row = {}
                    cols = st.columns(len(st.session_state.df.columns))
                    for j, col_name in enumerate(st.session_state.df.columns):
                        with cols[j]:
                            if col_name in ['Job_ID', 'Processing_Time', 'Due_Date', 'Weight', 'Release_Date']:
                                edited_row[col_name] = st.number_input(
                                    f"{col_name}",
                                    value=float(row[col_name]),
                                    step=1.0,
                                    key=f"edit_{i}_{col_name}"
                                )
                            else:
                                edited_row[col_name] = st.text_input(
                                    f"{col_name}",
                                    value=str(row[col_name]),
                                    key=f"edit_{i}_{col_name}"
                                )

                    if st.button(f"Update Row {i+1}", key=f"update_{i}"):
                        for col_name, val in edited_row.items():
                            st.session_state.df.at[i, col_name] = val
                        st.success(f"Row {i+1} updated!")

        elif edit_option == "Add new rows":
            st.write("Add a new row:")
            new_row = {}
            cols = st.columns(len(st.session_state.df.columns))
            for j, col_name in enumerate(st.session_state.df.columns):
                with cols[j]:
                    if col_name in ['Job_ID', 'Processing_Time', 'Due_Date', 'Weight', 'Release_Date']:
                        default_val = st.session_state.df[col_name].max() + 1 if col_name == 'Job_ID' else 1
                        new_row[col_name] = st.number_input(
                            f"{col_name}",
                            value=float(default_val),
                            step=1.0,
                            key=f"new_{col_name}"
                        )
                    else:
                        new_row[col_name] = st.text_input(
                            f"{col_name}",
                            value="",
                            key=f"new_{col_name}"
                        )

            if st.button("Add Row"):
                # Append the new row to the session state dataframe
                st.session_state.df = pd.concat([st.session_state.df, pd.DataFrame([new_row])], ignore_index=True)
                st.success("New row added!")

        elif edit_option == "Delete rows":
            st.write("Select rows to delete:")
            rows_to_delete = []
            for i, row in st.session_state.df.iterrows():
                col1, col2 = st.columns([1, 3])
                with col1:
                    delete_this_row = st.checkbox(f"Delete row {i+1}", key=f"delete_{i}")
                    if delete_this_row:
                        rows_to_delete.append(i)
                with col2:
                    st.write(row.to_dict())

            if rows_to_delete and st.button("Delete Selected Rows"):
                st.session_state.df = st.session_state.df.drop(rows_to_delete).reset_index(drop=True)
                st.success(f"{len(rows_to_delete)} row(s) deleted!")

        # Display the updated data
        st.subheader("Current Data (after any edits)")
        st.dataframe(st.session_state.df)

        # Add a button to clear the data if needed
        if st.button("Clear uploaded data and start over"):
            st.session_state.df = None
            st.session_state.file_uploaded = False
            st.rerun()

        # Make sure we use the session state dataframe for algorithms
        df = st.session_state.df.copy()  # Important! Make a copy for the algorithms
else:
    # Manual job input
    st.subheader("Enter job details")

    # If we already have manually entered data, use it
    if 'df' in st.session_state and st.session_state.df is not None and not st.session_state.file_uploaded:
        num_jobs = st.number_input("Number of jobs", min_value=1, max_value=20, value=len(st.session_state.df), key="manual_num_jobs")

        # If jobs were added or removed, update the dataframe
        if num_jobs != len(st.session_state.df):
            # If increasing, add new rows
            if num_jobs > len(st.session_state.df):
                for i in range(len(st.session_state.df), num_jobs):
                    new_row = {
                        "Job_ID": i+1,
                        "Processing_Time": 5,
                        "Due_Date": 10,
                        "Weight": 1,
                        "Release_Date": 0
                    }
                    st.session_state.df = pd.concat([st.session_state.df, pd.DataFrame([new_row])], ignore_index=True)
            # If decreasing, remove rows
            else:
                st.session_state.df = st.session_state.df.iloc[:num_jobs].reset_index(drop=True)

        # Show editable fields for each job
        job_data = {
            "Job_ID": [],
            "Processing_Time": [],
            "Due_Date": [],
            "Weight": [],
            "Release_Date": []
        }

        for i in range(num_jobs):
            st.markdown(f"**Job {i+1}**")
            cols = st.columns(5)

            # Get existing values or defaults
            current_job = st.session_state.df.iloc[i] if i < len(st.session_state.df) else None

            job_data["Job_ID"].append(cols[0].number_input(
                f"Job ID {i+1}",
                value=int(current_job["Job_ID"]) if current_job is not None else i+1,
                key=f"id_{i}"
            ))
            job_data["Processing_Time"].append(cols[1].number_input(
                f"Processing Time {i+1}",
                value=int(current_job["Processing_Time"]) if current_job is not None else 5,
                min_value=1,
                key=f"pt_{i}"
            ))
            job_data["Due_Date"].append(cols[2].number_input(
                f"Due Date {i+1}",
                value=int(current_job["Due_Date"]) if current_job is not None else 10,
                min_value=1,
                key=f"dd_{i}"
            ))
            job_data["Weight"].append(cols[3].number_input(
                f"Weight {i+1}",
                value=int(current_job["Weight"]) if current_job is not None else 1,
                min_value=1,
                key=f"w_{i}"
            ))
            job_data["Release_Date"].append(cols[4].number_input(
                f"Release Date {i+1}",
                value=int(current_job["Release_Date"]) if current_job is not None else 0,
                min_value=0,
                key=f"rd_{i}"
            ))

        # Update the session state dataframe
        st.session_state.df = pd.DataFrame(job_data)
        st.session_state.file_uploaded = False  # This is manual data, not uploaded

    else:
        # Starting fresh with manual input
        num_jobs = st.number_input("Number of jobs", min_value=1, max_value=20, value=3, key="manual_num_jobs")

        job_data = {
            "Job_ID": [],
            "Processing_Time": [],
            "Due_Date": [],
            "Weight": [],
            "Release_Date": []
        }

        for i in range(num_jobs):
            st.markdown(f"**Job {i+1}**")
            cols = st.columns(5)

            job_data["Job_ID"].append(cols[0].number_input(f"Job ID {i+1}", value=i+1, key=f"id_{i}"))
            job_data["Processing_Time"].append(cols[1].number_input(f"Processing Time {i+1}", value=5, min_value=1, key=f"pt_{i}"))
            job_data["Due_Date"].append(cols[2].number_input(f"Due Date {i+1}", value=10, min_value=1, key=f"dd_{i}"))
            job_data["Weight"].append(cols[3].number_input(f"Weight {i+1}", value=1, min_value=1, key=f"w_{i}"))
            job_data["Release_Date"].append(cols[4].number_input(f"Release Date {i+1}", value=0, min_value=0, key=f"rd_{i}"))

        # Create dataframe and store in session state
        st.session_state.df = pd.DataFrame(job_data)
        st.session_state.file_uploaded = False  # This is manual data, not uploaded

    # Display the current data
    st.write("Entered data:")
    st.dataframe(st.session_state.df)

    # Set df for algorithms
    df = st.session_state.df.copy()

# Helper function for Excel download with multiple sheets and embedded image
def get_excel_download_link(results_df, summary_df, additional_df, fig, algorithm_name):
    """Creates a downloadable Excel file with multiple sheets containing all results and the Gantt chart"""
    # Create a BytesIO object to store the Excel file
    output = io.BytesIO()

    # Create Excel writer
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Add each dataframe as a separate sheet
        results_df.to_excel(writer, sheet_name='Job Metrics', index=False)
        summary_df.to_excel(writer, sheet_name='Summary Statistics', index=False)
        additional_df.to_excel(writer, sheet_name='Additional Metrics', index=False)

        # Add an empty sheet for the Gantt chart
        workbook = writer.book
        gantt_sheet = workbook.create_sheet(title='Gantt Chart')

        # Add a title to the Gantt chart sheet
        gantt_sheet['A1'] = f"{algorithm_name} Schedule Gantt Chart"

        # Apply styling to the title using proper openpyxl styles
        gantt_sheet['A1'].font = Font(size=14, bold=True)
        gantt_sheet['A1'].alignment = Alignment(horizontal='center')

        # Merge cells for the title
        gantt_sheet.merge_cells('A1:G1')

        # Save the figure as a PNG in memory
        img_buf = io.BytesIO()
        fig.savefig(img_buf, format='png', dpi=150, bbox_inches='tight')
        img_buf.seek(0)

        # Create a PIL Image object from the buffer
        pil_img = PIL.Image.open(img_buf)

        # Create an openpyxl image
        xl_img = XLImage(img_buf)

        # Resize the image if needed (optional)
        scale_factor = 0.8  # Adjust this for sizing
        xl_img.width = int(pil_img.width * scale_factor)
        xl_img.height = int(pil_img.height * scale_factor)

        # Add the image to the sheet
        gantt_sheet.add_image(xl_img, 'A3')

        # Auto-adjust columns width for all sheets
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            for column in ws.columns:
                column_letter = get_column_letter(column[0].column)
                if column_letter == 'A':  # Adjust first column which often has headers
                    ws.column_dimensions[column_letter].width = 20
                else:
                    ws.column_dimensions[column_letter].width = 15

    # Set buffer position to start
    output.seek(0)

    # Convert to base64 for download link
    b64 = base64.b64encode(output.getvalue()).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{algorithm_name}_results.xlsx">📊 Download All Results (Excel)</a>'

    return href

# Run the algorithm
if df is not None and st.button("Run Algorithm"):
    # Check if Release_Date column exists, add if not
    if 'Release_Date' not in df.columns:
        df['Release_Date'] = 0  # Default to zero if not provided

    # Check if Weight column exists, add if not
    if 'Weight' not in df.columns:
        df['Weight'] = 1  # Default to one if not provided

    # 1. FCFS Algorithm
    if algorithm == "First Come First Serve (FCFS)":
        # Create a placeholder for the algorithm result
        result_container = st.container()
        figure_container = st.container()
        download_container = st.container()

        # Define FCFS function
        def FCFS(jobs_df):
            # Make a copy (no sorting - take jobs in the order they appear)
            df = jobs_df.copy()

            # Initialize variables
            current_time = 0
            start_times = []
            completion_times = []
            waiting_times = []
            flow_times = []
            lateness_values = []
            tardiness_values = []

            # Process each job in the order they appear
            for idx, job in df.iterrows():
                # Update current time (job can't start before its release date)
                current_time = max(current_time, job['Release_Date'])

                # Calculate start time for this job
                start_time = current_time
                start_times.append(start_time)

                # Process the job
                current_time += job['Processing_Time']
                completion_time = current_time
                completion_times.append(completion_time)

                # Calculate metrics
                waiting_time = start_time - job['Release_Date']
                flow_time = completion_time - job['Release_Date']
                lateness = completion_time - job['Due_Date']
                tardiness = max(0, completion_time - job['Due_Date'])

                # Save results
                waiting_times.append(waiting_time)
                flow_times.append(flow_time)
                lateness_values.append(lateness)
                tardiness_values.append(tardiness)

            # Add results to the dataframe
            df['Start_Time'] = start_times
            df['Completion_Time'] = completion_times
            df['Wait_Time'] = waiting_times
            df['Flow_Time'] = flow_times
            df['Lateness'] = lateness_values
            df['Tardiness'] = tardiness_values

            # Calculate total and average metrics
            total_completion_time = sum(completion_times)
            avg_completion_time = round(total_completion_time / len(jobs_df), 2)
            total_waiting_time = sum(waiting_times)
            avg_waiting_time = round(total_waiting_time / len(jobs_df), 2)
            total_flow_time = sum(flow_times)
            avg_flow_time = round(total_flow_time / len(jobs_df), 2)
            total_lateness = sum(lateness_values)
            avg_lateness = round(total_lateness / len(jobs_df), 2)
            total_tardiness = sum(tardiness_values)
            avg_tardiness = round(total_tardiness / len(jobs_df), 2)
            max_tardiness = max(tardiness_values)
            makespan = max(completion_times)

            # Find minimum values
            min_completion_time = min(completion_times)
            min_waiting_time = min(waiting_times)
            min_flow_time = min(flow_times)
            min_lateness = min(lateness_values)
            min_tardiness = min(tardiness_values)

            # Calculate other metrics
            total_processing_time = sum(df['Processing_Time'])
            utilization = round((total_processing_time / makespan) * 100, 2) if makespan > 0 else 0
            avg_jobs_in_system = round(total_flow_time / makespan, 2) if makespan > 0 else 0
            num_tardy_jobs = sum(1 for t in tardiness_values if t > 0)

            # Weighted metrics
            weighted_completion_times = [c * w for c, w in zip(completion_times, df['Weight'])]
            weighted_flow_times = [f * w for f, w in zip(flow_times, df['Weight'])]
            weighted_wait_times = [wt * w for wt, w in zip(waiting_times, df['Weight'])]
            weighted_tardiness = [t * w for t, w in zip(tardiness_values, df['Weight'])]

            sum_weighted_completion = sum(weighted_completion_times)
            sum_weighted_flow = sum(weighted_flow_times)
            sum_weighted_wait = sum(weighted_wait_times)
            sum_weighted_tardiness = sum(weighted_tardiness)

            # Display results in Streamlit
            with result_container:
                st.subheader("Results: FCFS")
                st.write(f"Job sequence: {df['Job_ID'].tolist()}")
                st.write(f"Makespan (total completion time): {makespan}")

                # Create a results table for all individual job metrics
                results_table = df[['Job_ID','Start_Time', 'Completion_Time', 'Wait_Time',
                                  'Flow_Time', 'Lateness', 'Tardiness']]
                st.subheader("Detailed Job Metrics:")
                st.dataframe(results_table)

                # Create a summary table with averages
                summary_data = {
                    'Metric': ['Completion Time', 'Wait Time', 'Flow Time', 'Lateness', 'Tardiness'],
                    'Total': [total_completion_time, total_waiting_time, total_flow_time, total_lateness, total_tardiness],
                    'Minimum': [min_completion_time, min_waiting_time, min_flow_time, min_lateness, min_tardiness],
                    'Average': [avg_completion_time, avg_waiting_time, avg_flow_time, avg_lateness, avg_tardiness],
                    'Maximum': [max(completion_times), max(waiting_times), max(flow_times), max(lateness_values), max_tardiness]
                }
                summary_table = pd.DataFrame(summary_data)
                st.subheader("Summary Statistics:")
                st.dataframe(summary_table)

                # Create additional metrics table
                additional_metrics = {
                    'Metric': ['Utilization (%)', 'Avg Jobs in System', 'Number of Tardy Jobs',
                              'Weighted Completion Time', 'Weighted Flow Time', 'Weighted Wait Time', 'Weighted Tardiness'],
                    'Value': [utilization, avg_jobs_in_system, num_tardy_jobs,
                             sum_weighted_completion, sum_weighted_flow, sum_weighted_wait, sum_weighted_tardiness]
                }
                additional_table = pd.DataFrame(additional_metrics)
                st.subheader("Additional Metrics:")
                st.dataframe(additional_table)

            # Create visualization
            fig, ax = plt.subplots(figsize=(20, 5))
            visualize_fcfs_schedule(df, fig, ax)

            with figure_container:
                st.pyplot(fig)

            # Add download links
            with download_container:
                st.subheader("Download Results")

                # Excel download with all data
                st.markdown(
                    get_excel_download_link(results_table, summary_table, additional_table, fig, "FCFS"),
                    unsafe_allow_html=True
                )

            return df, results_table, summary_table, additional_table, fig

        def visualize_fcfs_schedule(df, fig, ax):
            """Visualize the FCFS schedule with a Gantt chart"""
            # Define colors for jobs
            colors = ['red', 'orange', 'blue', 'green', 'purple', 'brown', 'pink', 'gray', 'cyan', 'magenta']

            y_position = 0
            y_ticks = []
            y_labels = []

            # Draw jobs in FCFS order
            for idx, job in df.iterrows():
                job_id = int(job['Job_ID'])
                release_time = job['Release_Date']
                start_time = job['Start_Time']
                duration = job['Processing_Time']
                completion_time = job['Completion_Time']
                due_date = job['Due_Date']

                # Choose color based on job index
                color = colors[job_id % len(colors)]

                # Draw waiting time (lighter shade)
                if start_time > release_time:
                    waiting_rect = patches.Rectangle(
                        (release_time, y_position),
                        start_time - release_time,
                        0.6,
                        linewidth=1,
                        edgecolor='black',
                        facecolor=color,
                        alpha=0.3
                    )
                    ax.add_patch(waiting_rect)
                    ax.text(
                        release_time + (start_time - release_time)/2,
                        y_position + 0.3,
                        'Wait',
                        ha='center',
                        va='center',
                        fontsize=8
                    )

                # Draw processing time (full color)
                process_rect = patches.Rectangle(
                    (start_time, y_position),
                    duration,
                    0.6,
                    linewidth=1,
                    edgecolor='black',
                    facecolor=color
                )
                ax.add_patch(process_rect)

                # Add job ID text in the center of the rectangle
                ax.text(
                    start_time + duration/2,
                    y_position + 0.3,
                    f"Job {job_id}",
                    color='white',
                    fontweight='bold',
                    ha='center',
                    va='center'
                )

                # Add time markers
                ax.annotate(f"R:{int(release_time)}", (release_time, y_position-0.3), ha='center', fontsize=8)
                ax.annotate(f"D:{int(due_date)}", (completion_time + 0.5, y_position + 0.2), ha='center', fontsize=8, color='green')
                ax.annotate(f"S:{int(start_time)}", (start_time, y_position-0.3), ha='center', fontsize=8)
                ax.annotate(f"C:{int(completion_time)}", (completion_time, y_position-0.3), ha='center', fontsize=8)

                # Next job position
                y_position += 1
                y_ticks.append(y_position - 0.7)
                y_labels.append(f"Job {job_id}")

            # Set up the axis
            max_time = max(df['Completion_Time'].max(), df['Due_Date'].max()) + 2
            ax.set_xlim(-1, max_time)
            ax.set_ylim(-0.5, y_position)

            # Add time markers
            time_ticks = np.arange(0, max_time + 5, 5)
            ax.set_xticks(time_ticks)
            ax.set_yticks(y_ticks)
            ax.set_yticklabels(y_labels)

            # Add grid
            ax.grid(True, axis='x', linestyle='-', alpha=0.3)

            # Add labels and title
            ax.set_xlabel('Time')
            ax.set_title('FCFS (First Come First Serve) Schedule')

            plt.tight_layout()

        # Run the algorithm
        FCFS(df)

    # 2. SPT Algorithm
    elif algorithm == "Shortest Processing Time (SPT)":
        # Create a placeholder for the algorithm result
        result_container = st.container()
        figure_container = st.container()
        download_container = st.container()

        def SPT(jobs_df):
            # Make a copy (sorting by the increasing processing time)
            df = jobs_df.copy().sort_values('Processing_Time')

            # Initialize variables
            current_time = 0
            start_times = []
            completion_times = []
            waiting_times = []
            flow_times = []
            lateness_values = []
            tardiness_values = []

            # Process each job in the order they appear
            for idx, job in df.iterrows():
                # Update current time (job can't start before its release date)
                current_time = max(current_time, job['Release_Date'])

                # Calculate start time for this job
                start_time = current_time
                start_times.append(start_time)

                # Process the job
                current_time += job['Processing_Time']
                completion_time = current_time
                completion_times.append(completion_time)

                # Calculate metrics
                waiting_time = start_time - job['Release_Date']
                flow_time = completion_time - job['Release_Date']
                lateness = completion_time - job['Due_Date']
                tardiness = max(0, completion_time - job['Due_Date'])

                # Save results
                waiting_times.append(waiting_time)
                flow_times.append(flow_time)
                lateness_values.append(lateness)
                tardiness_values.append(tardiness)

            # Add results to the dataframe
            df['Start_Time'] = start_times
            df['Completion_Time'] = completion_times
            df['Wait_Time'] = waiting_times
            df['Flow_Time'] = flow_times
            df['Lateness'] = lateness_values
            df['Tardiness'] = tardiness_values

            # Calculate total and average metrics
            total_completion_time = sum(completion_times)
            avg_completion_time = round(total_completion_time / len(jobs_df), 2)
            total_waiting_time = sum(waiting_times)
            avg_waiting_time = round(total_waiting_time / len(jobs_df), 2)
            total_flow_time = sum(flow_times)
            avg_flow_time = round(total_flow_time / len(jobs_df), 2)
            total_lateness = sum(lateness_values)
            avg_lateness = round(total_lateness / len(jobs_df), 2)
            total_tardiness = sum(tardiness_values)
            avg_tardiness = round(total_tardiness / len(jobs_df), 2)
            max_tardiness = max(tardiness_values)
            makespan = max(completion_times)
            
            # Find minimum values
            min_completion_time = min(completion_times)
            min_waiting_time = min(waiting_times)
            min_flow_time = min(flow_times)
            min_lateness = min(lateness_values)
            min_tardiness = min(tardiness_values)

            # Calculate other metrics
            total_processing_time = sum(df['Processing_Time'])
            utilization = round((total_processing_time / total_flow_time) * 100, 2) if total_flow_time > 0 else 0
            avg_jobs_in_system = round(total_flow_time / total_processing_time, 2) if total_processing_time > 0 else 0
            num_tardy_jobs = sum(1 for t in tardiness_values if t > 0)

            # Weighted metrics
            weighted_completion_times = [c * w for c, w in zip(completion_times, df['Weight'])]
            weighted_flow_times = [f * w for f, w in zip(flow_times, df['Weight'])]
            weighted_wait_times = [wt * w for wt, w in zip(waiting_times, df['Weight'])]
            weighted_tardiness = [t * w for t, w in zip(tardiness_values, df['Weight'])]

            sum_weighted_completion = sum(weighted_completion_times)
            sum_weighted_flow = sum(weighted_flow_times)
            sum_weighted_wait = sum(weighted_wait_times)
            sum_weighted_tardiness = sum(weighted_tardiness)

            # Display results in Streamlit
            with result_container:
                st.subheader("Results: SPT")
                st.write(f"Job sequence: {df['Job_ID'].tolist()}")
                st.write(f"Makespan (total completion time): {makespan}")

                # Create a results table for job metrics
                results_table = df[['Job_ID','Start_Time', 'Completion_Time', 'Wait_Time',
                                  'Flow_Time', 'Lateness', 'Tardiness']]
                st.subheader("Detailed Job Metrics:")
                st.dataframe(results_table)

                # Create a summary table with averages
                summary_data = {
                    'Metric': ['Completion Time', 'Wait Time', 'Flow Time', 'Lateness', 'Tardiness'],
                    'Total': [total_completion_time, total_waiting_time, total_flow_time, total_lateness, total_tardiness],
                    'Minimum': [min_completion_time, min_waiting_time, min_flow_time, min_lateness, min_tardiness],
                    'Average': [avg_completion_time, avg_waiting_time, avg_flow_time, avg_lateness, avg_tardiness],
                    'Maximum': [max(completion_times), max(waiting_times), max(flow_times), max(lateness_values), max_tardiness]
                }
                summary_table = pd.DataFrame(summary_data)
                st.subheader("Summary Statistics:")
                st.dataframe(summary_table)

                # Create additional metrics table
                additional_metrics = {
                    'Metric': ['Utilization (%)', 'Avg Jobs in System', 'Number of Tardy Jobs',
                              'Weighted Completion Time', 'Weighted Flow Time', 'Weighted Wait Time', 'Weighted Tardiness'],
                    'Value': [utilization, avg_jobs_in_system, num_tardy_jobs,
                             sum_weighted_completion, sum_weighted_flow, sum_weighted_wait, sum_weighted_tardiness]
                }
                additional_table = pd.DataFrame(additional_metrics)
                st.subheader("Additional Metrics:")
                st.dataframe(additional_table)

            # Create visualization
            fig, ax = plt.subplots(figsize=(20, 5))
            visualize_spt_schedule(df, fig, ax)
            with figure_container:
                st.pyplot(fig)
            with download_container:
                st.subheader("Download Results")

                # Excel download with all data
                st.markdown(
                    get_excel_download_link(results_table, summary_table, additional_table, fig, "SPT"),
                    unsafe_allow_html=True
                )

            return df, results_table, summary_table, additional_table, fig

        def visualize_spt_schedule(df, fig, ax):
            """Visualize the SPT schedule with a Gantt chart"""
            # Define colors for jobs
            colors = ['red', 'orange', 'blue', 'green', 'purple', 'brown', 'pink', 'gray', 'cyan', 'magenta']

            y_position = 0
            y_ticks = []
            y_labels = []

            # Draw jobs in SPT order
            for idx, job in df.iterrows():
                job_id = int(job['Job_ID'])
                release_time = job['Release_Date']
                start_time = job['Start_Time']
                duration = job['Processing_Time']
                completion_time = job['Completion_Time']
                due_date = job['Due_Date']

                # Choose color based on job index
                color = colors[job_id % len(colors)]

                # Draw waiting time (lighter shade)
                if start_time > release_time:
                    waiting_rect = patches.Rectangle(
                        (release_time, y_position),
                        start_time - release_time,
                        0.6,
                        linewidth=1,
                        edgecolor='black',
                        facecolor=color,
                        alpha=0.3
                    )
                    ax.add_patch(waiting_rect)
                    ax.text(
                        release_time + (start_time - release_time)/2,
                        y_position + 0.3,
                        'Wait',
                        ha='center',
                        va='center',
                        fontsize=8
                    )

                # Draw processing time (full color)
                process_rect = patches.Rectangle(
                    (start_time, y_position),
                    duration,
                    0.6,
                    linewidth=1,
                    edgecolor='black',
                    facecolor=color
                )
                ax.add_patch(process_rect)

                # Add job ID text in the center of the rectangle
                ax.text(
                    start_time + duration/2,
                    y_position + 0.3,
                    f"Job {job_id}",
                    color='white',
                    fontweight='bold',
                    ha='center',
                    va='center'
                )

                # Add time markers
                ax.annotate(f"R:{int(release_time)}", (release_time, y_position-0.3), ha='center', fontsize=8)
                ax.annotate(f"D:{int(due_date)}", (completion_time + 0.5, y_position + 0.2), ha='center', fontsize=8, color='green')
                ax.annotate(f"S:{int(start_time)}", (start_time, y_position-0.3), ha='center', fontsize=8)
                ax.annotate(f"C:{int(completion_time)}", (completion_time, y_position-0.3), ha='center', fontsize=8)

                # Next job position
                y_position += 1
                y_ticks.append(y_position - 0.7)
                y_labels.append(f"Job {job_id}")

            # Set up the axis
            max_time = max(df['Completion_Time'].max(), df['Due_Date'].max()) + 2
            ax.set_xlim(-1, max_time)
            ax.set_ylim(-0.5, y_position)

            # Add time markers
            time_ticks = np.arange(0, max_time + 5, 5)
            ax.set_xticks(time_ticks)
            ax.set_yticks(y_ticks)
            ax.set_yticklabels(y_labels)

            # Add grid
            ax.grid(True, axis='x', linestyle='-', alpha=0.3)

            # Add labels and title
            ax.set_xlabel('Time')
            ax.set_title('SPT (Shortest Processing Time) Schedule')

            plt.tight_layout()

        # Run the algorithm
        SPT(df)

    # 3. LPT Algorithm
    elif algorithm == "Longest Processing Time (LPT)":
        # Create a placeholder for the algorithm result
        result_container = st.container()
        figure_container = st.container()
        download_container = st.container()

        def LPT(jobs_df):
            # Make a copy (sorting by the decreasing processing time)
            df = jobs_df.copy().sort_values('Processing_Time', ascending=False)

            # Initialize variables
            current_time = 0
            start_times = []
            completion_times = []
            waiting_times = []
            flow_times = []
            lateness_values = []
            tardiness_values = []

            # Process each job in the order they appear
            for idx, job in df.iterrows():
                # Update current time (job can't start before its release date)
                current_time = max(current_time, job['Release_Date'])

                # Calculate start time for this job
                start_time = current_time
                start_times.append(start_time)

                # Process the job
                current_time += job['Processing_Time']
                completion_time = current_time
                completion_times.append(completion_time)

                # Calculate metrics
                waiting_time = start_time - job['Release_Date']
                flow_time = completion_time - job['Release_Date']
                lateness = completion_time - job['Due_Date']
                tardiness = max(0, completion_time - job['Due_Date'])

                # Save results
                waiting_times.append(waiting_time)
                flow_times.append(flow_time)
                lateness_values.append(lateness)
                tardiness_values.append(tardiness)

            # Add results to the dataframe
            df['Start_Time'] = start_times
            df['Completion_Time'] = completion_times
            df['Wait_Time'] = waiting_times
            df['Flow_Time'] = flow_times
            df['Lateness'] = lateness_values
            df['Tardiness'] = tardiness_values

            # Calculate total and average metrics
            total_completion_time = sum(completion_times)
            avg_completion_time = round(total_completion_time / len(jobs_df), 2)
            total_waiting_time = sum(waiting_times)
            avg_waiting_time = round(total_waiting_time / len(jobs_df), 2)
            total_flow_time = sum(flow_times)
            avg_flow_time = round(total_flow_time / len(jobs_df), 2)
            total_lateness = sum(lateness_values)
            avg_lateness = round(total_lateness / len(jobs_df), 2)
            total_tardiness = sum(tardiness_values)
            avg_tardiness = round(total_tardiness / len(jobs_df), 2)
            max_tardiness = max(tardiness_values)
            makespan = max(completion_times)

            # Find minimum values
            min_completion_time = min(completion_times)
            min_waiting_time = min(waiting_times)
            min_flow_time = min(flow_times)
            min_lateness = min(lateness_values)
            min_tardiness = min(tardiness_values)

            # Calculate other metrics
            total_processing_time = sum(df['Processing_Time'])
            utilization = round((total_processing_time / total_flow_time) * 100, 2) if total_flow_time > 0 else 0
            avg_jobs_in_system = round(total_flow_time / total_processing_time, 2) if total_processing_time > 0 else 0
            num_tardy_jobs = sum(1 for t in tardiness_values if t > 0)

            # Weighted metrics
            weighted_completion_times = [c * w for c, w in zip(completion_times, df['Weight'])]
            weighted_flow_times = [f * w for f, w in zip(flow_times, df['Weight'])]
            weighted_wait_times = [wt * w for wt, w in zip(waiting_times, df['Weight'])]
            weighted_tardiness = [t * w for t, w in zip(tardiness_values, df['Weight'])]

            sum_weighted_completion = sum(weighted_completion_times)
            sum_weighted_flow = sum(weighted_flow_times)
            sum_weighted_wait = sum(weighted_wait_times)
            sum_weighted_tardiness = sum(weighted_tardiness)

            # Display results in Streamlit
            with result_container:
                st.subheader("Results: LPT")
                st.write(f"Job sequence: {df['Job_ID'].tolist()}")
                st.write(f"Makespan (total completion time): {makespan}")

                # Create a results table for job metrics
                results_table = df[['Job_ID','Start_Time', 'Completion_Time', 'Wait_Time',
                                  'Flow_Time', 'Lateness', 'Tardiness']]
                st.subheader("Detailed Job Metrics:")
                st.dataframe(results_table)

                # Create a summary table with averages
                summary_data = {
                    'Metric': ['Completion Time', 'Wait Time', 'Flow Time', 'Lateness', 'Tardiness'],
                    'Total': [total_completion_time, total_waiting_time, total_flow_time, total_lateness, total_tardiness],
                    'Minimum': [min_completion_time, min_waiting_time, min_flow_time, min_lateness, min_tardiness],
                    'Average': [avg_completion_time, avg_waiting_time, avg_flow_time, avg_lateness, avg_tardiness],
                    'Maximum': [max(completion_times), max(waiting_times), max(flow_times), max(lateness_values), max_tardiness]
                }
                summary_table = pd.DataFrame(summary_data)
                st.subheader("Summary Statistics:")
                st.dataframe(summary_table)

                # Create additional metrics table
                additional_metrics = {
                    'Metric': ['Utilization (%)', 'Avg Jobs in System', 'Number of Tardy Jobs',
                              'Weighted Completion Time', 'Weighted Flow Time', 'Weighted Wait Time', 'Weighted Tardiness'],
                    'Value': [utilization, avg_jobs_in_system, num_tardy_jobs,
                             sum_weighted_completion, sum_weighted_flow, sum_weighted_wait, sum_weighted_tardiness]
                }
                additional_table = pd.DataFrame
                additional_table = pd.DataFrame(additional_metrics)
                st.subheader("Additional Metrics:")
                st.dataframe(additional_table)

            # Create visualization
            fig, ax = plt.subplots(figsize=(20, 5))
            visualize_lpt_schedule(df, fig, ax)
            with figure_container:
                st.pyplot(fig)
            with download_container:
                st.subheader("Download Results")

                # Excel download with all data
                st.markdown(
                    get_excel_download_link(results_table, summary_table, additional_table, fig, "LPT"),
                    unsafe_allow_html=True
                )

            return df, results_table, summary_table, additional_table, fig


        def visualize_lpt_schedule(df, fig, ax):
            """Visualize the LPT schedule with a Gantt chart"""
            # Define colors for jobs
            colors = ['red', 'orange', 'blue', 'green', 'purple', 'brown', 'pink', 'gray', 'cyan', 'magenta']

            y_position = 0
            y_ticks = []
            y_labels = []

            # Draw jobs in LPT order
            for idx, job in df.iterrows():
                job_id = int(job['Job_ID'])
                release_time = job['Release_Date']
                start_time = job['Start_Time']
                duration = job['Processing_Time']
                completion_time = job['Completion_Time']
                due_date = job['Due_Date']

                # Choose color based on job index
                color = colors[job_id % len(colors)]

                # Draw waiting time (lighter shade)
                if start_time > release_time:
                    waiting_rect = patches.Rectangle(
                        (release_time, y_position),
                        start_time - release_time,
                        0.6,
                        linewidth=1,
                        edgecolor='black',
                        facecolor=color,
                        alpha=0.3
                    )
                    ax.add_patch(waiting_rect)
                    ax.text(
                        release_time + (start_time - release_time)/2,
                        y_position + 0.3,
                        'Wait',
                        ha='center',
                        va='center',
                        fontsize=8
                    )

                # Draw processing time (full color)
                process_rect = patches.Rectangle(
                    (start_time, y_position),
                    duration,
                    0.6,
                    linewidth=1,
                    edgecolor='black',
                    facecolor=color
                )
                ax.add_patch(process_rect)

                # Add job ID text in the center of the rectangle
                ax.text(
                    start_time + duration/2,
                    y_position + 0.3,
                    f"Job {job_id}",
                    color='white',
                    fontweight='bold',
                    ha='center',
                    va='center'
                )

                # Add time markers
                ax.annotate(f"R:{int(release_time)}", (release_time, y_position-0.3), ha='center', fontsize=8)
                ax.annotate(f"D:{int(due_date)}", (completion_time + 0.5, y_position + 0.2), ha='center', fontsize=8, color='green')
                ax.annotate(f"S:{int(start_time)}", (start_time, y_position-0.3), ha='center', fontsize=8)
                ax.annotate(f"C:{int(completion_time)}", (completion_time, y_position-0.3), ha='center', fontsize=8)

                # Next job position
                y_position += 1
                y_ticks.append(y_position - 0.7)
                y_labels.append(f"Job {job_id}")

            # Set up the axis
            max_time = max(df['Completion_Time'].max(), df['Due_Date'].max()) + 2
            ax.set_xlim(-1, max_time)
            ax.set_ylim(-0.5, y_position)

            # Add time markers
            time_ticks = np.arange(0, max_time + 5, 5)
            ax.set_xticks(time_ticks)
            ax.set_yticks(y_ticks)
            ax.set_yticklabels(y_labels)

            # Add grid
            ax.grid(True, axis='x', linestyle='-', alpha=0.3)

            # Add labels and title
            ax.set_xlabel('Time')
            ax.set_title('LPT (Longest Processing Time) Schedule')

            plt.tight_layout()

        # Run the algorithm
        LPT(df)

    # 4. EDD Algorithm
    elif algorithm == "Earliest Due Date (EDD)":
        # Create a placeholder for the algorithm result
        result_container = st.container()
        figure_container = st.container()
        download_container = st.container()

        def EDD(jobs_df):
            # Make a copy (sorting by the increasing due_date)
            df = jobs_df.copy().sort_values('Due_Date')

            # Initialize variables
            current_time = 0
            start_times = []
            completion_times = []
            waiting_times = []
            flow_times = []
            lateness_values = []
            tardiness_values = []

            # Process each job in the order they appear
            for idx, job in df.iterrows():
                # Update current time (job can't start before its release date)
                current_time = max(current_time, job['Release_Date'])

                # Calculate start time for this job
                start_time = current_time
                start_times.append(start_time)

                # Process the job
                current_time += job['Processing_Time']
                completion_time = current_time
                completion_times.append(completion_time)

                # Calculate metrics
                waiting_time = start_time - job['Release_Date']
                flow_time = completion_time - job['Release_Date']
                lateness = completion_time - job['Due_Date']
                tardiness = max(0, completion_time - job['Due_Date'])

                # Save results
                waiting_times.append(waiting_time)
                flow_times.append(flow_time)
                lateness_values.append(lateness)
                tardiness_values.append(tardiness)

            # Add results to the dataframe
            df['Start_Time'] = start_times
            df['Completion_Time'] = completion_times
            df['Wait_Time'] = waiting_times
            df['Flow_Time'] = flow_times
            df['Lateness'] = lateness_values
            df['Tardiness'] = tardiness_values

            # Calculate total and average metrics
            total_completion_time = sum(completion_times)
            avg_completion_time = round(total_completion_time / len(jobs_df), 2)
            total_waiting_time = sum(waiting_times)
            avg_waiting_time = round(total_waiting_time / len(jobs_df), 2)
            total_flow_time = sum(flow_times)
            avg_flow_time = round(total_flow_time / len(jobs_df), 2)
            total_lateness = sum(lateness_values)
            avg_lateness = round(total_lateness / len(jobs_df), 2)
            total_tardiness = sum(tardiness_values)
            avg_tardiness = round(total_tardiness / len(jobs_df), 2)
            max_tardiness = max(tardiness_values)
            makespan = max(completion_times)

            # Find minimum values
            min_completion_time = min(completion_times)
            min_waiting_time = min(waiting_times)
            min_flow_time = min(flow_times)
            min_lateness = min(lateness_values)
            min_tardiness = min(tardiness_values)

            # Calculate other metrics
            total_processing_time = sum(df['Processing_Time'])
            utilization = round((total_processing_time / total_flow_time) * 100, 2) if total_flow_time > 0 else 0
            avg_jobs_in_system = round(total_flow_time / total_processing_time, 2) if total_processing_time > 0 else 0
            num_tardy_jobs = sum(1 for t in tardiness_values if t > 0)

            # Weighted metrics
            weighted_completion_times = [c * w for c, w in zip(completion_times, df['Weight'])]
            weighted_flow_times = [f * w for f, w in zip(flow_times, df['Weight'])]
            weighted_wait_times = [wt * w for wt, w in zip(waiting_times, df['Weight'])]
            weighted_tardiness = [t * w for t, w in zip(tardiness_values, df['Weight'])]

            sum_weighted_completion = sum(weighted_completion_times)
            sum_weighted_flow = sum(weighted_flow_times)
            sum_weighted_wait = sum(weighted_wait_times)
            sum_weighted_tardiness = sum(weighted_tardiness)

            # Display results in Streamlit
            with result_container:
                st.subheader("Results: EDD")
                st.write(f"Job sequence: {df['Job_ID'].tolist()}")
                st.write(f"Makespan (total completion time): {makespan}")

                # Create a results table for job metrics
                results_table = df[['Job_ID','Start_Time', 'Completion_Time', 'Wait_Time',
                                  'Flow_Time', 'Lateness', 'Tardiness']]
                st.subheader("Detailed Job Metrics:")
                st.dataframe(results_table)

                # Create a summary table with averages
                summary_data = {
                    'Metric': ['Completion Time', 'Wait Time', 'Flow Time', 'Lateness', 'Tardiness'],
                    'Total': [total_completion_time, total_waiting_time, total_flow_time, total_lateness, total_tardiness],
                    'Minimum': [min_completion_time, min_waiting_time, min_flow_time, min_lateness, min_tardiness],
                    'Average': [avg_completion_time, avg_waiting_time, avg_flow_time, avg_lateness, avg_tardiness],
                    'Maximum': [max(completion_times), max(waiting_times), max(flow_times), max(lateness_values), max_tardiness]
                }
                summary_table = pd.DataFrame(summary_data)
                st.subheader("Summary Statistics:")
                st.dataframe(summary_table)

                # Create additional metrics table
                additional_metrics = {
                    'Metric': ['Utilization (%)', 'Avg Jobs in System', 'Number of Tardy Jobs',
                              'Weighted Completion Time', 'Weighted Flow Time', 'Weighted Wait Time', 'Weighted Tardiness'],
                    'Value': [utilization, avg_jobs_in_system, num_tardy_jobs,
                             sum_weighted_completion, sum_weighted_flow, sum_weighted_wait, sum_weighted_tardiness]
                }
                additional_table = pd.DataFrame(additional_metrics)
                st.subheader("Additional Metrics:")
                st.dataframe(additional_table)

            # Create visualization
            fig, ax = plt.subplots(figsize=(20, 5))
            visualize_edd_schedule(df, fig, ax)
            with figure_container:
                st.pyplot(fig)
            with download_container:
                st.subheader("Download Results")

                # Excel download with all data
                st.markdown(
                    get_excel_download_link(results_table, summary_table, additional_table, fig, "EDD"),
                    unsafe_allow_html=True
                )

            return df, results_table, summary_table, additional_table, fig


        def visualize_edd_schedule(df, fig, ax):
            """Visualize the EDD schedule with a Gantt chart"""
            # Define colors for jobs
            colors = ['red', 'orange', 'blue', 'green', 'purple', 'brown', 'pink', 'gray', 'cyan', 'magenta']

            y_position = 0
            y_ticks = []
            y_labels = []

            # Draw jobs in EDD order
            for idx, job in df.iterrows():
                job_id = int(job['Job_ID'])
                release_time = job['Release_Date']
                start_time = job['Start_Time']
                duration = job['Processing_Time']
                completion_time = job['Completion_Time']
                due_date = job['Due_Date']

                # Choose color based on job index
                color = colors[job_id % len(colors)]

                # Draw waiting time (lighter shade)
                if start_time > release_time:
                    waiting_rect = patches.Rectangle(
                        (release_time, y_position),
                        start_time - release_time,
                        0.6,
                        linewidth=1,
                        edgecolor='black',
                        facecolor=color,
                        alpha=0.3
                    )
                    ax.add_patch(waiting_rect)
                    ax.text(
                        release_time + (start_time - release_time)/2,
                        y_position + 0.3,
                        'Wait',
                        ha='center',
                        va='center',
                        fontsize=8
                    )

                # Draw processing time (full color)
                process_rect = patches.Rectangle(
                    (start_time, y_position),
                    duration,
                    0.6,
                    linewidth=1,
                    edgecolor='black',
                    facecolor=color
                )
                ax.add_patch(process_rect)

                # Add job ID text in the center of the rectangle
                ax.text(
                    start_time + duration/2,
                    y_position + 0.3,
                    f"Job {job_id}",
                    color='white',
                    fontweight='bold',
                    ha='center',
                    va='center'
                )

                # Add time markers
                ax.annotate(f"R:{int(release_time)}", (release_time, y_position-0.3), ha='center', fontsize=8)
                ax.annotate(f"D:{int(due_date)}", (completion_time + 0.5, y_position + 0.2), ha='center', fontsize=8, color='green')
                ax.annotate(f"S:{int(start_time)}", (start_time, y_position-0.3), ha='center', fontsize=8)
                ax.annotate(f"C:{int(completion_time)}", (completion_time, y_position-0.3), ha='center', fontsize=8)

                # Next job position
                y_position += 1
                y_ticks.append(y_position - 0.7)
                y_labels.append(f"Job {job_id}")

            # Set up the axis
            max_time = max(df['Completion_Time'].max(), df['Due_Date'].max()) + 2
            ax.set_xlim(-1, max_time)
            ax.set_ylim(-0.5, y_position)

            # Add time markers
            time_ticks = np.arange(0, max_time + 5, 5)
            ax.set_xticks(time_ticks)
            ax.set_yticks(y_ticks)
            ax.set_yticklabels(y_labels)

            # Add grid
            ax.grid(True, axis='x', linestyle='-', alpha=0.3)

            # Add labels and title
            ax.set_xlabel('Time')
            ax.set_title('EDD (Earliest Due Date) Schedule')

            plt.tight_layout()

        # Run the algorithm
        EDD(df)

    # 5. ERD Algorithm
    elif algorithm == "Earliest Release Date (ERD)":
        # Create a placeholder for the algorithm result
        result_container = st.container()
        figure_container = st.container()
        download_container = st.container()

        def ERD(jobs_df):
            # Make a copy (sorting by the increasing release_date)
            df = jobs_df.copy().sort_values('Release_Date')

            # Initialize variables
            current_time = 0
            start_times = []
            completion_times = []
            waiting_times = []
            flow_times = []
            lateness_values = []
            tardiness_values = []

            # Process each job in the order they appear
            for idx, job in df.iterrows():
                # Update current time (job can't start before its release date)
                current_time = max(current_time, job['Release_Date'])

                # Calculate start time for this job
                start_time = current_time
                start_times.append(start_time)

                # Process the job
                current_time += job['Processing_Time']
                completion_time = current_time
                completion_times.append(completion_time)

                # Calculate metrics
                waiting_time = start_time - job['Release_Date']
                flow_time = completion_time - job['Release_Date']
                lateness = completion_time - job['Due_Date']
                tardiness = max(0, completion_time - job['Due_Date'])

                # Save results
                waiting_times.append(waiting_time)
                flow_times.append(flow_time)
                lateness_values.append(lateness)
                tardiness_values.append(tardiness)

            # Add results to the dataframe
            df['Start_Time'] = start_times
            df['Completion_Time'] = completion_times
            df['Wait_Time'] = waiting_times
            df['Flow_Time'] = flow_times
            df['Lateness'] = lateness_values
            df['Tardiness'] = tardiness_values

            # Calculate total and average metrics
            total_completion_time = sum(completion_times)
            avg_completion_time = round(total_completion_time / len(jobs_df), 2)
            total_waiting_time = sum(waiting_times)
            avg_waiting_time = round(total_waiting_time / len(jobs_df), 2)
            total_flow_time = sum(flow_times)
            avg_flow_time = round(total_flow_time / len(jobs_df), 2)
            total_lateness = sum(lateness_values)
            avg_lateness = round(total_lateness / len(jobs_df), 2)
            total_tardiness = sum(tardiness_values)
            avg_tardiness = round(total_tardiness / len(jobs_df), 2)
            max_tardiness = max(tardiness_values)
            makespan = max(completion_times)

            # Find minimum values
            min_completion_time = min(completion_times)
            min_waiting_time = min(waiting_times)
            min_flow_time = min(flow_times)
            min_lateness = min(lateness_values)
            min_tardiness = min(tardiness_values)

            # Calculate other metrics
            total_processing_time = sum(df['Processing_Time'])
            utilization = round((total_processing_time / total_flow_time) * 100, 2) if total_flow_time > 0 else 0
            avg_jobs_in_system = round(total_flow_time / total_processing_time, 2) if total_processing_time > 0 else 0
            num_tardy_jobs = sum(1 for t in tardiness_values if t > 0)

            # Weighted metrics
            weighted_completion_times = [c * w for c, w in zip(completion_times, df['Weight'])]
            weighted_flow_times = [f * w for f, w in zip(flow_times, df['Weight'])]
            weighted_wait_times = [wt * w for wt, w in zip(waiting_times, df['Weight'])]
            weighted_tardiness = [t * w for t, w in zip(tardiness_values, df['Weight'])]

            sum_weighted_completion = sum(weighted_completion_times)
            sum_weighted_flow = sum(weighted_flow_times)
            sum_weighted_wait = sum(weighted_wait_times)
            sum_weighted_tardiness = sum(weighted_tardiness)

            # Display results in Streamlit
            with result_container:
                st.subheader("Results: ERD")
                st.write(f"Job sequence: {df['Job_ID'].tolist()}")
                st.write(f"Makespan (total completion time): {makespan}")

                # Create a results table for job metrics
                results_table = df[['Job_ID','Start_Time', 'Completion_Time', 'Wait_Time',
                                  'Flow_Time', 'Lateness', 'Tardiness']]
                st.subheader("Detailed Job Metrics:")
                st.dataframe(results_table)

                # Create a summary table with averages
                summary_data = {
                    'Metric': ['Completion Time', 'Wait Time', 'Flow Time', 'Lateness', 'Tardiness'],
                    'Total': [total_completion_time, total_waiting_time, total_flow_time, total_lateness, total_tardiness],
                    'Minimum': [min_completion_time, min_waiting_time, min_flow_time, min_lateness, min_tardiness],
                    'Average': [avg_completion_time, avg_waiting_time, avg_flow_time, avg_lateness, avg_tardiness],
                    'Maximum': [max(completion_times), max(waiting_times), max(flow_times), max(lateness_values), max_tardiness]
                }
                summary_table = pd.DataFrame(summary_data)
                st.subheader("Summary Statistics:")
                st.dataframe(summary_table)

                # Create additional metrics table
                additional_metrics = {
                    'Metric': ['Utilization (%)', 'Avg Jobs in System', 'Number of Tardy Jobs',
                              'Weighted Completion Time', 'Weighted Flow Time', 'Weighted Wait Time', 'Weighted Tardiness'],
                    'Value': [utilization, avg_jobs_in_system, num_tardy_jobs,
                             sum_weighted_completion, sum_weighted_flow, sum_weighted_wait, sum_weighted_tardiness]
                }
                additional_table = pd.DataFrame(additional_metrics)
                st.subheader("Additional Metrics:")
                st.dataframe(additional_table)

            # Create visualization
            fig, ax = plt.subplots(figsize=(20, 5))
            visualize_erd_schedule(df, fig, ax)
            with figure_container:
                st.pyplot(fig)
            with download_container:
                st.subheader("Download Results")

                # Excel download with all data
                st.markdown(
                    get_excel_download_link(results_table, summary_table, additional_table, fig, "ERD"),
                    unsafe_allow_html=True
                )

            return df, results_table, summary_table, additional_table, fig

        def visualize_erd_schedule(df, fig, ax):
            """Visualize the ERD schedule with a Gantt chart"""
            # Define colors for jobs
            colors = ['red', 'orange', 'blue', 'green', 'purple', 'brown', 'pink', 'gray', 'cyan', 'magenta']

            y_position = 0
            y_ticks = []
            y_labels = []

            # Draw jobs in ERD order
            for idx, job in df.iterrows():
                job_id = int(job['Job_ID'])
                release_time = job['Release_Date']
                start_time = job['Start_Time']
                duration = job['Processing_Time']
                completion_time = job['Completion_Time']
                due_date = job['Due_Date']

                # Choose color based on job index
                color = colors[job_id % len(colors)]

                # Draw waiting time (lighter shade)
                if start_time > release_time:
                    waiting_rect = patches.Rectangle(
                        (release_time, y_position),
                        start_time - release_time,
                        0.6,
                        linewidth=1,
                        edgecolor='black',
                        facecolor=color,
                        alpha=0.3
                    )
                    ax.add_patch(waiting_rect)
                    ax.text(
                        release_time + (start_time - release_time)/2,
                        y_position + 0.3,
                        'Wait',
                        ha='center',
                        va='center',
                        fontsize=8
                    )

                # Draw processing time (full color)
                process_rect = patches.Rectangle(
                    (start_time, y_position),
                    duration,
                    0.6,
                    linewidth=1,
                    edgecolor='black',
                    facecolor=color
                )
                ax.add_patch(process_rect)

                # Add job ID text in the center of the rectangle
                ax.text(
                    start_time + duration/2,
                    y_position + 0.3,
                    f"Job {job_id}",
                    color='white',
                    fontweight='bold',
                    ha='center',
                    va='center'
                )

                # Add time markers
                ax.annotate(f"R:{int(release_time)}", (release_time, y_position-0.3), ha='center', fontsize=8)
                ax.annotate(f"D:{int(due_date)}", (completion_time + 0.5, y_position + 0.2), ha='center', fontsize=8, color='green')
                ax.annotate(f"S:{int(start_time)}", (start_time, y_position-0.3), ha='center', fontsize=8)
                ax.annotate(f"C:{int(completion_time)}", (completion_time, y_position-0.3), ha='center', fontsize=8)

                # Next job position
                y_position += 1
                y_ticks.append(y_position - 0.7)
                y_labels.append(f"Job {job_id}")

            # Set up the axis
            max_time = max(df['Completion_Time'].max(), df['Due_Date'].max()) + 2
            ax.set_xlim(-1, max_time)
            ax.set_ylim(-0.5, y_position)

            # Add time markers
            time_ticks = np.arange(0, max_time + 5, 5)
            ax.set_xticks(time_ticks)
            ax.set_yticks(y_ticks)
            ax.set_yticklabels(y_labels)

            # Add grid
            ax.grid(True, axis='x', linestyle='-', alpha=0.3)

            # Add labels and title
            ax.set_xlabel('Time')
            ax.set_title('ERD (Earliest Release Date) Schedule')

            plt.tight_layout()

        # Run the algorithm
        ERD(df)

    # 6. WSPT Algorithm
    elif algorithm == "Weighted Shortest Processing Time (WSPT)":
        # Create a placeholder for the algorithm result
        result_container = st.container()
        figure_container = st.container()
        download_container = st.container()

        def WSPT(jobs_df):
            # Make a copy and calculate pj/wj ratio
            df = jobs_df.copy()
            df['pj/wj'] = round(df['Processing_Time']/df['Weight'], 2)

            # Sort by the increasing pj/wj ratio
            df = df.sort_values('pj/wj')

            # Initialize variables
            current_time = 0
            start_times = []
            completion_times = []
            waiting_times = []
            flow_times = []
            lateness_values = []
            tardiness_values = []

            # Process each job in the order they appear
            for idx, job in df.iterrows():
                # Update current time (job can't start before its release date)
                current_time = max(current_time, job['Release_Date'])

                # Calculate start time for this job
                start_time = current_time
                start_times.append(start_time)

                # Process the job
                current_time += job['Processing_Time']
                completion_time = current_time
                completion_times.append(completion_time)

                # Calculate metrics
                waiting_time = start_time - job['Release_Date']
                flow_time = completion_time - job['Release_Date']
                lateness = completion_time - job['Due_Date']
                tardiness = max(0, completion_time - job['Due_Date'])

                # Save results
                waiting_times.append(waiting_time)
                flow_times.append(flow_time)
                lateness_values.append(lateness)
                tardiness_values.append(tardiness)

            # Add results to the dataframe
            df['Start_Time'] = start_times
            df['Completion_Time'] = completion_times
            df['Wait_Time'] = waiting_times
            df['Flow_Time'] = flow_times
            df['Lateness'] = lateness_values
            df['Tardiness'] = tardiness_values

            # Calculate total and average metrics
            total_completion_time = sum(completion_times)
            avg_completion_time = round(total_completion_time / len(jobs_df), 2)
            total_waiting_time = sum(waiting_times)
            avg_waiting_time = round(total_waiting_time / len(jobs_df), 2)
            total_flow_time = sum(flow_times)
            avg_flow_time = round(total_flow_time / len(jobs_df), 2)
            total_lateness = sum(lateness_values)
            avg_lateness = round(total_lateness / len(jobs_df), 2)
            total_tardiness = sum(tardiness_values)
            avg_tardiness = round(total_tardiness / len(jobs_df), 2)
            max_tardiness = max(tardiness_values)
            makespan = max(completion_times)

            # Find minimum values
            min_completion_time = min(completion_times)
            min_waiting_time = min(waiting_times)
            min_flow_time = min(flow_times)
            min_lateness = min(lateness_values)
            min_tardiness = min(tardiness_values)

            # Calculate other metrics
            total_processing_time = sum(df['Processing_Time'])
            utilization = round((total_processing_time / total_flow_time) * 100, 2) if total_flow_time > 0 else 0
            avg_jobs_in_system = round(total_flow_time / total_processing_time, 2) if total_processing_time > 0 else 0
            num_tardy_jobs = sum(1 for t in tardiness_values if t > 0)

            # Weighted metrics
            weighted_completion_times = [c * w for c, w in zip(completion_times, df['Weight'])]
            weighted_flow_times = [f * w for f, w in zip(flow_times, df['Weight'])]
            weighted_wait_times = [wt * w for wt, w in zip(waiting_times, df['Weight'])]
            weighted_tardiness = [t * w for t, w in zip(tardiness_values, df['Weight'])]

            sum_weighted_completion = sum(weighted_completion_times)
            sum_weighted_flow = sum(weighted_flow_times)
            sum_weighted_wait = sum(weighted_wait_times)
            sum_weighted_tardiness = sum(weighted_tardiness)

            # Display results in Streamlit
            with result_container:
                st.subheader("Results: WSPT")
                st.write(f"Job sequence: {df['Job_ID'].tolist()}")
                st.write(f"Makespan (total completion time): {makespan}")

                # Create a results table for job metrics
                results_table = df[['Job_ID', 'pj/wj','Start_Time', 'Completion_Time', 'Wait_Time',
                                  'Flow_Time', 'Lateness', 'Tardiness']]
                st.subheader("Detailed Job Metrics:")
                st.dataframe(results_table)

                # Create a summary table with averages
                summary_data = {
                    'Metric': ['Completion Time', 'Wait Time', 'Flow Time', 'Lateness', 'Tardiness'],
                    'Total': [total_completion_time, total_waiting_time, total_flow_time, total_lateness, total_tardiness],
                    'Minimum': [min_completion_time, min_waiting_time, min_flow_time, min_lateness, min_tardiness],
                    'Average': [avg_completion_time, avg_waiting_time, avg_flow_time, avg_lateness, avg_tardiness],
                    'Maximum': [max(completion_times), max(waiting_times), max(flow_times), max(lateness_values), max_tardiness]
                }
                summary_table = pd.DataFrame(summary_data)
                st.subheader("Summary Statistics:")
                st.dataframe(summary_table)

                # Create additional metrics table
                additional_metrics = {
                    'Metric': ['Utilization (%)', 'Avg Jobs in System', 'Number of Tardy Jobs',
                              'Weighted Completion Time', 'Weighted Flow Time', 'Weighted Wait Time', 'Weighted Tardiness'],
                    'Value': [utilization, avg_jobs_in_system, num_tardy_jobs,
                             sum_weighted_completion, sum_weighted_flow, sum_weighted_wait, sum_weighted_tardiness]
                }
                additional_table = pd.DataFrame(additional_metrics)
                st.subheader("Additional Metrics:")
                st.dataframe(additional_table)

            # Create visualization
            fig, ax = plt.subplots(figsize=(20, 5))
            visualize_wspt_schedule(df, fig, ax)
            with figure_container:
                st.pyplot(fig)
            with download_container:
                st.subheader("Download Results")

                # Excel download with all data
                st.markdown(
                    get_excel_download_link(results_table, summary_table, additional_table, fig, "WSPT"),
                    unsafe_allow_html=True
                )

            return df, results_table, summary_table, additional_table, fig

        def visualize_wspt_schedule(df, fig, ax):
            """Visualize the WSPT schedule with a Gantt chart"""
            # Define colors for jobs
            colors = ['red', 'orange', 'blue', 'green', 'purple', 'brown', 'pink', 'gray', 'cyan', 'magenta']

            y_position = 0
            y_ticks = []
            y_labels = []

            # Draw jobs in WSPT order
            for idx, job in df.iterrows():
                job_id = int(job['Job_ID'])
                release_time = job['Release_Date']
                start_time = job['Start_Time']
                duration = job['Processing_Time']
                completion_time = job['Completion_Time']
                due_date = job['Due_Date']

                # Choose color based on job index
                color = colors[job_id % len(colors)]

                # Draw waiting time (lighter shade)
                if start_time > release_time:
                    waiting_rect = patches.Rectangle(
                        (release_time, y_position),
                        start_time - release_time,
                        0.6,
                        linewidth=1,
                        edgecolor='black',
                        facecolor=color,
                        alpha=0.3
                    )
                    ax.add_patch(waiting_rect)
                    ax.text(
                        release_time + (start_time - release_time)/2,
                        y_position + 0.3,
                        'Wait',
                        ha='center',
                        va='center',
                        fontsize=8
                    )

                # Draw processing time (full color)
                process_rect = patches.Rectangle(
                    (start_time, y_position),
                    duration,
                    0.6,
                    linewidth=1,
                    edgecolor='black',
                    facecolor=color
                )
                ax.add_patch(process_rect)

                # Add job ID text in the center of the rectangle
                ax.text(
                    start_time + duration/2,
                    y_position + 0.3,
                    f"Job {job_id}",
                    color='white',
                    fontweight='bold',
                    ha='center',
                    va='center'
                )

                # Add time markers
                ax.annotate(f"R:{int(release_time)}", (release_time, y_position-0.3), ha='center', fontsize=8)
                ax.annotate(f"D:{int(due_date)}", (completion_time + 0.5, y_position + 0.2), ha='center', fontsize=8, color='green')
                ax.annotate(f"S:{int(start_time)}", (start_time, y_position-0.3), ha='center', fontsize=8)
                ax.annotate(f"C:{int(completion_time)}", (completion_time, y_position-0.3), ha='center', fontsize=8)

                # Next job position
                y_position += 1
                y_ticks.append(y_position - 0.7)
                y_labels.append(f"Job {job_id}")

            # Set up the axis
            max_time = max(df['Completion_Time'].max(), df['Due_Date'].max()) + 2
            ax.set_xlim(-1, max_time)
            ax.set_ylim(-0.5, y_position)

            # Add time markers
            time_ticks = np.arange(0, max_time + 5, 5)
            ax.set_xticks(time_ticks)
            ax.set_yticks(y_ticks)
            ax.set_yticklabels(y_labels)

            # Add grid
            ax.grid(True, axis='x', linestyle='-', alpha=0.3)

            # Add labels and title
            ax.set_xlabel('Time')
            ax.set_title('WSPT (Weighted Shortest Processing Time) Schedule')

            plt.tight_layout()

        # Run the algorithm
        WSPT(df)

    # 7. Random Sequencing Algorithm
    elif algorithm == "Random Sequencing (RAND)":
        # Create a placeholder for the algorithm result
        result_container = st.container()
        figure_container = st.container()
        download_container = st.container()

        def RAND(jobs_df):
            # Make a copy and randomly shuffle the jobs
            df = jobs_df.sample(frac=1).reset_index(drop=True)

            # Initialize variables
            current_time = 0
            start_times = []
            completion_times = []
            waiting_times = []
            flow_times = []
            lateness_values = []
            tardiness_values = []

            # Process each job in the order they appear
            for idx, job in df.iterrows():
                # Update current time (job can't start before its release date)
                current_time = max(current_time, job['Release_Date'])

                # Calculate start time for this job
                start_time = current_time
                start_times.append(start_time)

                # Process the job
                current_time += job['Processing_Time']
                completion_time = current_time
                completion_times.append(completion_time)

                # Calculate metrics
                waiting_time = start_time - job['Release_Date']
                flow_time = completion_time - job['Release_Date']
                lateness = completion_time - job['Due_Date']
                tardiness = max(0, completion_time - job['Due_Date'])

                # Save results
                waiting_times.append(waiting_time)
                flow_times.append(flow_time)
                lateness_values.append(lateness)
                tardiness_values.append(tardiness)

            # Add results to the dataframe
            df['Start_Time'] = start_times
            df['Completion_Time'] = completion_times
            df['Wait_Time'] = waiting_times
            df['Flow_Time'] = flow_times
            df['Lateness'] = lateness_values
            df['Tardiness'] = tardiness_values

            # Calculate total and average metrics
            total_completion_time = sum(completion_times)
            avg_completion_time = round(total_completion_time / len(jobs_df), 2)
            total_waiting_time = sum(waiting_times)
            avg_waiting_time = round(total_waiting_time / len(jobs_df), 2)
            total_flow_time = sum(flow_times)
            avg_flow_time = round(total_flow_time / len(jobs_df), 2)
            total_lateness = sum(lateness_values)
            avg_lateness = round(total_lateness / len(jobs_df), 2)
            total_tardiness = sum(tardiness_values)
            avg_tardiness = round(total_tardiness / len(jobs_df), 2)
            max_tardiness = max(tardiness_values)
            makespan = max(completion_times)

            # Find minimum values
            min_completion_time = min(completion_times)
            min_waiting_time = min(waiting_times)
            min_flow_time = min(flow_times)
            min_lateness = min(lateness_values)
            min_tardiness = min(tardiness_values)

            # Calculate other metrics
            total_processing_time = sum(df['Processing_Time'])
            utilization = round((total_processing_time / total_flow_time) * 100, 2) if total_flow_time > 0 else 0
            avg_jobs_in_system = round(total_flow_time / total_processing_time, 2) if total_processing_time > 0 else 0
            num_tardy_jobs = sum(1 for t in tardiness_values if t > 0)

            # Weighted metrics
            weighted_completion_times = [c * w for c, w in zip(completion_times, df['Weight'])]
            weighted_flow_times = [f * w for f, w in zip(flow_times, df['Weight'])]
            weighted_wait_times = [wt * w for wt, w in zip(waiting_times, df['Weight'])]
            weighted_tardiness = [t * w for t, w in zip(tardiness_values, df['Weight'])]

            sum_weighted_completion = sum(weighted_completion_times)
            sum_weighted_flow = sum(weighted_flow_times)
            sum_weighted_wait = sum(weighted_wait_times)
            sum_weighted_tardiness = sum(weighted_tardiness)

            # Display results in Streamlit
            with result_container:
                st.subheader("Results: Random Sequencing")
                st.write(f"Job sequence: {df['Job_ID'].tolist()}")
                st.write(f"Makespan (total completion time): {makespan}")

                # Create a results table for job metrics
                results_table = df[['Job_ID','Start_Time', 'Completion_Time', 'Wait_Time',
                                  'Flow_Time', 'Lateness', 'Tardiness']]
                st.subheader("Detailed Job Metrics:")
                st.dataframe(results_table)

                # Create a summary table with averages
                summary_data = {
                    'Metric': ['Completion Time', 'Wait Time', 'Flow Time', 'Lateness', 'Tardiness'],
                    'Total': [total_completion_time, total_waiting_time, total_flow_time, total_lateness, total_tardiness],
                    'Minimum': [min_completion_time, min_waiting_time, min_flow_time, min_lateness, min_tardiness],
                    'Average': [avg_completion_time, avg_waiting_time, avg_flow_time, avg_lateness, avg_tardiness],
                    'Maximum': [max(completion_times), max(waiting_times), max(flow_times), max(lateness_values), max_tardiness]
                }
                summary_table = pd.DataFrame(summary_data)
                st.subheader("Summary Statistics:")
                st.dataframe(summary_table)

                # Create additional metrics table
                additional_metrics = {
                    'Metric': ['Utilization (%)', 'Avg Jobs in System', 'Number of Tardy Jobs',
                              'Weighted Completion Time', 'Weighted Flow Time', 'Weighted Wait Time', 'Weighted Tardiness'],
                    'Value': [utilization, avg_jobs_in_system, num_tardy_jobs,
                             sum_weighted_completion, sum_weighted_flow, sum_weighted_wait, sum_weighted_tardiness]
                }
                additional_table = pd.DataFrame(additional_metrics)
                st.subheader("Additional Metrics:")
                st.dataframe(additional_table)

            # Create visualization
            fig, ax = plt.subplots(figsize=(20, 5))
            visualize_rand_schedule(df, fig, ax)
            with figure_container:
                st.pyplot(fig)
            with download_container:
                st.subheader("Download Results")

                # Excel download with all data
                st.markdown(
                    get_excel_download_link(results_table, summary_table, additional_table, fig, "RAND"),
                    unsafe_allow_html=True
                )

            return df, results_table, summary_table, additional_table, fig

        def visualize_rand_schedule(df, fig, ax):
            """Visualize the RAND schedule with a Gantt chart"""
            # Define colors for jobs
            colors = ['red', 'orange', 'blue', 'green', 'purple', 'brown', 'pink', 'gray', 'cyan', 'magenta']

            y_position = 0
            y_ticks = []
            y_labels = []

            # Draw jobs in RAND order
            for idx, job in df.iterrows():
                job_id = int(job['Job_ID'])
                release_time = job['Release_Date']
                start_time = job['Start_Time']
                duration = job['Processing_Time']
                completion_time = job['Completion_Time']
                due_date = job['Due_Date']

                # Choose color based on job index
                color = colors[job_id % len(colors)]

                # Draw waiting time (lighter shade)
                if start_time > release_time:
                    waiting_rect = patches.Rectangle(
                        (release_time, y_position),
                        start_time - release_time,
                        0.6,
                        linewidth=1,
                        edgecolor='black',
                        facecolor=color,
                        alpha=0.3
                    )
                    ax.add_patch(waiting_rect)
                    ax.text(
                        release_time + (start_time - release_time)/2,
                        y_position + 0.3,
                        'Wait',
                        ha='center',
                        va='center',
                        fontsize=8
                    )

                # Draw processing time (full color)
                process_rect = patches.Rectangle(
                    (start_time, y_position),
                    duration,
                    0.6,
                    linewidth=1,
                    edgecolor='black',
                    facecolor=color
                )
                ax.add_patch(process_rect)

                # Add job ID text in the center of the rectangle
                ax.text(
                    start_time + duration/2,
                    y_position + 0.3,
                    f"Job {job_id}",
                    color='white',
                    fontweight='bold',
                    ha='center',
                    va='center'
                )

                # Add time markers
                ax.annotate(f"R:{int(release_time)}", (release_time, y_position-0.3), ha='center', fontsize=8)
                ax.annotate(f"D:{int(due_date)}", (completion_time + 0.5, y_position + 0.2), ha='center', fontsize=8, color='green')
                ax.annotate(f"S:{int(start_time)}", (start_time, y_position-0.3), ha='center', fontsize=8)
                ax.annotate(f"C:{int(completion_time)}", (completion_time, y_position-0.3), ha='center', fontsize=8)

                # Next job position
                y_position += 1
                y_ticks.append(y_position - 0.7)
                y_labels.append(f"Job {job_id}")

            # Set up the axis
            max_time = max(df['Completion_Time'].max(), df['Due_Date'].max()) + 2
            ax.set_xlim(-1, max_time)
            ax.set_ylim(-0.5, y_position)

            # Add time markers
            time_ticks = np.arange(0, max_time + 5, 5)
            ax.set_xticks(time_ticks)
            ax.set_yticks(y_ticks)
            ax.set_yticklabels(y_labels)

            # Add grid
            ax.grid(True, axis='x', linestyle='-', alpha=0.3)

            # Add labels and title
            ax.set_xlabel('Time')
            ax.set_title('RAND (Random Sequence) Schedule')

            plt.tight_layout()

        # Run the algorithm
        RAND(df)

    # 8. Moore's Rule Algorithm
    elif algorithm == "Moore's Rule (Minimize Late Jobs)":
        # Create a placeholder for the algorithm result
        result_container = st.container()
        figure_container = st.container()
        download_container = st.container()

        # Run Moore's algorithm
        def moore_rule(jobs_df):
            # Make a copy and sort by due date (ascending)
            df = jobs_df.copy().sort_values('Due_Date')

            # Initialize variables
            current_time = 0
            scheduled_indices = []

            # Process each task in order of due date
            for idx, task in df.iterrows():
                # Add this task to our schedule
                scheduled_indices.append(idx)
                current_time += task['Processing_Time']

                # If we've missed the due date
                if current_time > task['Due_Date']:
                    # Find and remove the task with the longest processing time
                    longest_idx = max(
                        [(i, df.loc[i, 'Processing_Time']) for i in scheduled_indices],
                        key=lambda x: x[1]
                    )[0]

                    scheduled_indices.remove(longest_idx)
                    current_time -= df.loc[longest_idx, 'Processing_Time']

            # Calculate late jobs
            all_indices = df.index.tolist()
            late_indices = [idx for idx in all_indices if idx not in scheduled_indices]

            # Get completed and late jobs
            completed_jobs = jobs_df.loc[scheduled_indices]
            late_jobs = jobs_df.loc[late_indices]

            # Calculate schedule and metrics
            completed_sequence = []
            current_time = 0
            start_times = {}
            completion_times = {}

            # Calculate schedule for completed jobs
            for idx, job in completed_jobs.sort_values('Due_Date').iterrows():
                job_id = int(job['Job_ID'])
                completed_sequence.append(job_id)

                start_times[job_id] = current_time
                current_time += job['Processing_Time']
                completion_times[job_id] = current_time

            # Store the end time of completed jobs to start scheduling late jobs
            completed_end_time = current_time

            # Calculate hypothetical schedule for late jobs
            late_sequence = []
            late_start_times = {}
            late_completion_times = {}

            # Continue from where completed jobs finished
            current_time = completed_end_time

            for idx, job in late_jobs.iterrows():
                job_id = int(job['Job_ID'])
                late_sequence.append(job_id)

                late_start_times[job_id] = current_time
                current_time += job['Processing_Time']
                late_completion_times[job_id] = current_time

            # Prepare result dataframes
            # 1. Completed jobs with schedule details
            completed_data = {
                'Job_ID': completed_sequence,
                'Processing_Time': [jobs_df[jobs_df['Job_ID'] == job_id]['Processing_Time'].values[0] for job_id in completed_sequence],
                'Due_Date': [jobs_df[jobs_df['Job_ID'] == job_id]['Due_Date'].values[0] for job_id in completed_sequence],
                'Start_Time': [start_times[job_id] for job_id in completed_sequence],
                'Completion_Time': [completion_times[job_id] for job_id in completed_sequence],
                'Status': ['On Time' for _ in completed_sequence]
            }
            completed_df = pd.DataFrame(completed_data)

            # 2. Late jobs
            late_job_ids = late_jobs['Job_ID'].tolist()
            late_data = {
                'Job_ID': late_job_ids,
                'Processing_Time': [jobs_df[jobs_df['Job_ID'] == job_id]['Processing_Time'].values[0] for job_id in late_job_ids],
                'Due_Date': [jobs_df[jobs_df['Job_ID'] == job_id]['Due_Date'].values[0] for job_id in late_job_ids],
                'Start_Time': [late_start_times[job_id] for job_id in late_job_ids],
                'Completion_Time': [late_completion_times[job_id] for job_id in late_job_ids],
                'Status': ['Late' for _ in late_job_ids]
            }
            late_df = pd.DataFrame(late_data)

            # 3. Summary metrics
            summary_data = {
                'Metric': ['Total Jobs', 'Jobs Completed On Time', 'Late Jobs', 'Objective Value'],
                'Value': [len(jobs_df), len(completed_jobs), len(late_jobs), len(late_jobs)]
            }
            summary_df = pd.DataFrame(summary_data)

            # Display results in Streamlit
            with result_container:
                st.subheader("Results: Moore's Rule")
                st.write(f"Number of jobs completed on time: {len(completed_jobs)} {completed_jobs['Job_ID'].tolist()}")
                st.write(f"Number of late jobs (Objective value): {len(late_jobs)} {late_jobs['Job_ID'].tolist()}")

                st.subheader("Completed Jobs:")
                st.dataframe(completed_df)

                st.subheader("Late Jobs:")
                st.dataframe(late_df)

                st.subheader("Summary Metrics:")
                st.dataframe(summary_df)

            # Create visualization
            fig, ax = plt.subplots(figsize=(20, 5))
            visualize_moore_schedule(jobs_df, scheduled_indices, late_indices, fig, ax)
            with figure_container:
                st.pyplot(fig)

            # Add download options
            with download_container:
                st.subheader("Download Results")

                # Excel download with all data
                st.markdown(
                    get_moore_excel_download_link(completed_df, late_df, summary_df, fig, "Moores_Rule"),
                    unsafe_allow_html=True
                )

            return completed_jobs, late_jobs, completed_df, late_df, summary_df, fig

        def get_moore_excel_download_link(completed_df, late_df, summary_df, fig, algorithm_name):
            """Creates a downloadable Excel file for Moore's Rule results"""
            # Create a BytesIO object to store the Excel file
            output = io.BytesIO()

            # Create Excel writer
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Add each dataframe as a separate sheet
                completed_df.to_excel(writer, sheet_name='Completed Jobs', index=False)
                late_df.to_excel(writer, sheet_name='Late Jobs', index=False)
                summary_df.to_excel(writer, sheet_name='Summary Metrics', index=False)

                # Add an empty sheet for the Gantt chart
                workbook = writer.book
                gantt_sheet = workbook.create_sheet(title='Schedule Visualization')

                # Add a title to the Gantt chart sheet
                gantt_sheet['A1'] = f"{algorithm_name} Schedule Visualization"

                # Apply styling to the title
                gantt_sheet['A1'].font = Font(size=14, bold=True)
                gantt_sheet['A1'].alignment = Alignment(horizontal='center')

                # Merge cells for the title
                gantt_sheet.merge_cells('A1:G1')

                # Save the figure as a PNG in memory
                img_buf = io.BytesIO()
                fig.savefig(img_buf, format='png', dpi=150, bbox_inches='tight')
                img_buf.seek(0)

                # Create a PIL Image object from the buffer
                pil_img = PIL.Image.open(img_buf)

                # Create an openpyxl image
                xl_img = XLImage(img_buf)

                # Resize the image if needed
                scale_factor = 0.8
                xl_img.width = int(pil_img.width * scale_factor)
                xl_img.height = int(pil_img.height * scale_factor)

                # Add the image to the sheet
                gantt_sheet.add_image(xl_img, 'A3')

                # Auto-adjust columns width for all sheets
                for sheet in workbook.sheetnames:
                    ws = workbook[sheet]
                    for column in ws.columns:
                        column_letter = get_column_letter(column[0].column)
                        if column_letter == 'A':
                            ws.column_dimensions[column_letter].width = 20
                        else:
                            ws.column_dimensions[column_letter].width = 15

            # Set buffer position to start
            output.seek(0)

            # Convert to base64 for download link
            b64 = base64.b64encode(output.getvalue()).decode()
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{algorithm_name}_results.xlsx">📊 Download All Results (Excel)</a>'

            return href

        def visualize_moore_schedule(jobs_df, completed_indices, late_indices, fig, ax):
            """Visualize the Moore schedule with completed and late jobs in a single row"""
            # Define colors for jobs
            colors = ['orange', 'green', 'brown', 'purple', 'gray', 'blue', 'pink', 'cyan', 'magenta', 'red']

            # Start with completed jobs
            completed_jobs = jobs_df.loc[completed_indices].sort_values('Due_Date')

            # Single row for all jobs
            y_position = 0

            # First draw completed jobs
            current_time = 0
            completed_start_time = current_time

            for idx, job in completed_jobs.iterrows():
                job_id = int(job['Job_ID'])
                duration = job['Processing_Time']

                # Choose color based on job index
                color = colors[job_id % len(colors)]

                # Create rectangle for the job
                rect = patches.Rectangle((current_time, y_position), duration, 1,
                                        linewidth=1, edgecolor='black', facecolor=color)
                ax.add_patch(rect)

                # Add job ID text in the center of the rectangle
                ax.text(current_time + duration/2, y_position + 0.5, f"Job {job_id}",
                        color='white', fontweight='bold', ha='center', va='center')

                # Move time forward
                current_time += duration

            # Save where completed jobs end
            completed_end_time = current_time

            # Draw a vertical line at the end of completed jobs
            ax.axvline(x=current_time, color='navy', linestyle='-', linewidth=2)

            # Place "Completed Jobs" label above completed jobs section
            ax.text((completed_start_time + completed_end_time) / 2, 1.3, "Completed Jobs",
                    fontsize=12, fontweight='bold', ha='center', va='center')

            # Now draw late jobs after the vertical line
            late_jobs = jobs_df.loc[late_indices]
            late_start_time = current_time

            for idx, job in late_jobs.iterrows():
                job_id = int(job['Job_ID'])
                duration = job['Processing_Time']

                # Choose color based on job index
                color = colors[job_id % len(colors)]

                # Create rectangle for the job
                rect = patches.Rectangle((current_time, y_position), duration, 1,
                                        linewidth=1, edgecolor='black', facecolor=color)
                ax.add_patch(rect)

                # Add job ID text in the center of the rectangle
                ax.text(current_time + duration/2, y_position + 0.5, f"Job {job_id}",
                        color='white', fontweight='bold', ha='center', va='center')

                # Move time forward
                current_time += duration

            # Save where late jobs end
            late_end_time = current_time

            # Place "Late Jobs" label above late jobs section
            if len(late_jobs) > 0:
                ax.text((late_start_time + late_end_time) / 2, 1.3, "Late Jobs",
                        fontsize=12, fontweight='bold', ha='center', va='center')

            # Set up the axis
            max_time = current_time
            ax.set_xlim(-1, max_time + 1)
            ax.set_ylim(-0.2, 1.5)  # Reduced height for single row

            # Add time markers
            time_ticks = np.arange(0, max_time + 4, 2)
            ax.set_xticks(time_ticks)

            # Remove y-axis ticks and labels
            ax.set_yticks([])
            ax.set_yticklabels([])

            # Add grid
            ax.grid(True, axis='x', linestyle='-', alpha=0.3)

            # Add x-axis label
            ax.set_xlabel('Time')

            plt.title('Moore Rule Schedule: Completed Jobs | Late Jobs')
            plt.tight_layout()

        # Run the algorithm
        moore_rule(df)

    # 9. SRPT Algorithm
    elif algorithm == "Shortest Remaining Processing Time (SRPT)":
        # Create a placeholder for the algorithm result
        result_container = st.container()
        figure_container = st.container()
        download_container = st.container()

        def SRPT(jobs_df):
            # Make a copy of the dataframe
            df = jobs_df.copy()

            # Create a column to track remaining processing time
            df['Remaining_Time'] = df['Processing_Time'].copy()

            # Sort by release date
            sorted_jobs = df.sort_values('Release_Date').reset_index(drop=True)

            # List to store execution timeline
            timeline = []

            # Keep track of current time and events
            current_time = sorted_jobs['Release_Date'].min()
            next_release_times = sorted_jobs['Release_Date'].tolist()
            next_release_times.sort()

            # Keep track of available and completed jobs
            available_jobs = []  # (job_id, remaining_time)
            completed_jobs = set()

            # Process until all jobs are completed
            while len(completed_jobs) < len(df):
                # Add any newly released jobs to available pool
                newly_released = sorted_jobs[(sorted_jobs['Release_Date'] <= current_time) &
                                            (~sorted_jobs['Job_ID'].isin([j[0] for j in available_jobs])) &
                                            (~sorted_jobs['Job_ID'].isin(completed_jobs))]

                for _, job in newly_released.iterrows():
                    available_jobs.append((job['Job_ID'], job['Remaining_Time']))

                # If no available jobs, jump to next release time
                if not available_jobs:
                    future_releases = [t for t in next_release_times if t > current_time]
                    if future_releases:
                        current_time = future_releases[0]
                        continue
                    else:
                        break  # No more jobs to process

                # Find job with shortest remaining time
                available_jobs.sort(key=lambda x: x[1])  # Sort by remaining time
                current_job_id, current_job_remaining = available_jobs[0]

                # Determine how long to run this job
                # Find the next event (job completion or new release)
                next_release = min([t for t in next_release_times if t > current_time], default=float('inf'))
                job_completion_time = current_time + current_job_remaining

                # The next event is either job completion or next release, whichever comes first
                next_event_time = min(job_completion_time, next_release)
                duration = next_event_time - current_time

                # Execute the job for this duration
                timeline.append({
                    'job_id': current_job_id,
                    'start': current_time,
                    'end': next_event_time
                })

                # Update remaining time
                job_index = sorted_jobs[sorted_jobs['Job_ID'] == current_job_id].index[0]
                sorted_jobs.loc[job_index, 'Remaining_Time'] -= duration

                # Update available jobs list
                if next_event_time == job_completion_time:
                    # Job completed
                    completed_jobs.add(current_job_id)
                    available_jobs.pop(0)  # Remove this job
                else:
                    # Job preempted, update its remaining time
                    available_jobs[0] = (current_job_id, current_job_remaining - duration)

                # Advance time
                current_time = next_event_time

            # Calculate job metrics for display
            job_metrics = {}
            for segment in timeline:
                job_id = segment['job_id']
                if job_id not in job_metrics:
                    job_metrics[job_id] = {
                        'start_time': float('inf'),
                        'completion_time': 0,
                        'total_processing_time': sorted_jobs[sorted_jobs['Job_ID'] == job_id]['Processing_Time'].values[0]
                    }

                # Update job metrics
                job_metrics[job_id]['start_time'] = min(job_metrics[job_id]['start_time'], segment['start'])
                job_metrics[job_id]['completion_time'] = max(job_metrics[job_id]['completion_time'], segment['end'])

            # Calculate additional metrics
            for job_id, metrics in job_metrics.items():
                job_row = sorted_jobs[sorted_jobs['Job_ID'] == job_id].iloc[0]
                metrics['release_time'] = job_row['Release_Date']
                metrics['due_date'] = job_row['Due_Date']
                metrics['flow_time'] = metrics['completion_time'] - metrics['release_time']
                metrics['lateness'] = metrics['completion_time'] - metrics['due_date']
                metrics['tardiness'] = max(0, metrics['lateness'])
                metrics['waiting_time'] = metrics['flow_time'] - metrics['total_processing_time']

            # Calculate preemption info
            preemptions = {}
            for job_id in job_metrics:
                segments = [seg for seg in timeline if seg['job_id'] == job_id]
                preemption_count = len(segments) - 1
                if preemption_count > 0:
                    preemptions[job_id] = preemption_count

            # Create dataframes for Excel export
            # 1. Timeline dataframe
            timeline_data = []
            for segment in timeline:
                timeline_data.append({
                    'Job_ID': segment['job_id'],
                    'Start_Time': segment['start'],
                    'End_Time': segment['end'],
                    'Duration': segment['end'] - segment['start']
                })
            timeline_df = pd.DataFrame(timeline_data)

            # 2. Job metrics dataframe
            job_metrics_data = []
            for job_id, metrics in job_metrics.items():
                job_metrics_data.append({
                    'Job_ID': job_id,
                    'Release_Time': metrics['release_time'],
                    'First_Start': metrics['start_time'],
                    'Completion_Time': metrics['completion_time'],
                    'Processing_Time': metrics['total_processing_time'],
                    'Flow_Time': metrics['flow_time'],
                    'Waiting_Time': metrics['waiting_time'],
                    'Due_Date': metrics['due_date'],
                    'Lateness': metrics['lateness'],
                    'Tardiness': metrics['tardiness']
                })
            job_metrics_df = pd.DataFrame(job_metrics_data)

            # 3. Preemption summary dataframe
            preemption_data = []
            for job_id, count in preemptions.items():
                preemption_data.append({
                    'Job_ID': job_id,
                    'Preemption_Count': count
                })
            preemption_df = pd.DataFrame(preemption_data) if preemption_data else pd.DataFrame({'Job_ID': [], 'Preemption_Count': []})

            # 4. Summary statistics
            makespan = max(segment['end'] for segment in timeline)
            total_flow_time = sum(metrics['flow_time'] for metrics in job_metrics.values())
            avg_flow_time = total_flow_time / len(job_metrics)
            total_waiting_time = sum(metrics['waiting_time'] for metrics in job_metrics.values())
            avg_waiting_time = total_waiting_time / len(job_metrics)
            total_tardiness = sum(metrics['tardiness'] for metrics in job_metrics.values())
            avg_tardiness = total_tardiness / len(job_metrics)

            summary_data = {
                'Metric': ['Makespan', 'Total Flow Time', 'Average Flow Time',
                          'Total Waiting Time', 'Average Waiting Time',
                          'Total Tardiness', 'Average Tardiness',
                          'Total Preemptions'],
                'Value': [makespan, total_flow_time, avg_flow_time,
                        total_waiting_time, avg_waiting_time,
                        total_tardiness, avg_tardiness,
                        sum(preemptions.values()) if preemptions else 0]
            }
            summary_df = pd.DataFrame(summary_data)

            # Display results in Streamlit
            with result_container:
                st.subheader("Results: SRPT")

                # Show timeline info
                st.write(f"Makespan (total completion time): {makespan}")

                # Show preemption info
                if preemptions:
                    st.subheader("Preemptions:")
                    for job_id, count in preemptions.items():
                        st.write(f"Job {job_id} was preempted {count} times")
                else:
                    st.write("No jobs were preempted")

                # Show job metrics
                st.subheader("Job Metrics:")
                st.dataframe(job_metrics_df)

                # Show summary statistics
                st.subheader("Summary Statistics:")
                st.dataframe(summary_df)

            # Visualize the timeline
            fig, ax = plt.subplots(figsize=(15, 5))
            visualize_srpt_timeline(timeline, sorted_jobs['Release_Date'].tolist(), fig, ax, show_release_lines)
            with figure_container:
                st.pyplot(fig)

            # Add download options
            with download_container:
                st.subheader("Download Results")

                # Excel download with all data
                st.markdown(
                    get_srpt_excel_download_link(timeline_df, job_metrics_df, preemption_df, summary_df, fig, "SRPT"),
                    unsafe_allow_html=True
                )

            return timeline, job_metrics, timeline_df, job_metrics_df, preemption_df, summary_df, fig

        def get_srpt_excel_download_link(timeline_df, job_metrics_df, preemption_df, summary_df, fig, algorithm_name):
            """Creates a downloadable Excel file for SRPT results"""
            # Create a BytesIO object to store the Excel file
            output = io.BytesIO()

            # Create Excel writer
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Add each dataframe as a separate sheet
                job_metrics_df.to_excel(writer, sheet_name='Job Metrics', index=False)
                timeline_df.to_excel(writer, sheet_name='Timeline Segments', index=False)
                preemption_df.to_excel(writer, sheet_name='Preemptions', index=False)
                summary_df.to_excel(writer, sheet_name='Summary Statistics', index=False)

                # Add an empty sheet for the Gantt chart
                workbook = writer.book
                gantt_sheet = workbook.create_sheet(title='Schedule Visualization')

                # Add a title to the Gantt chart sheet
                gantt_sheet['A1'] = f"{algorithm_name} Schedule Visualization"

                # Apply styling to the title
                gantt_sheet['A1'].font = Font(size=14, bold=True)
                gantt_sheet['A1'].alignment = Alignment(horizontal='center')

                # Merge cells for the title
                gantt_sheet.merge_cells('A1:G1')

                # Save the figure as a PNG in memory
                img_buf = io.BytesIO()
                fig.savefig(img_buf, format='png', dpi=150, bbox_inches='tight')
                img_buf.seek(0)

                # Create a PIL Image object from the buffer
                pil_img = PIL.Image.open(img_buf)

                # Create an openpyxl image
                xl_img = XLImage(img_buf)

                # Resize the image if needed
                scale_factor = 0.8
                xl_img.width = int(pil_img.width * scale_factor)
                xl_img.height = int(pil_img.height * scale_factor)

                # Add the image to the sheet
                gantt_sheet.add_image(xl_img, 'A3')

                # Auto-adjust columns width for all sheets
                for sheet in workbook.sheetnames:
                    ws = workbook[sheet]
                    for column in ws.columns:
                        column_letter = get_column_letter(column[0].column)
                        if column_letter == 'A':
                            ws.column_dimensions[column_letter].width = 20
                        else:
                            ws.column_dimensions[column_letter].width = 15

            # Set buffer position to start
            output.seek(0)

            # Convert to base64 for download link
            b64 = base64.b64encode(output.getvalue()).decode()
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{algorithm_name}_results.xlsx">📊 Download All Results (Excel)</a>'

            return href
        # 
        def visualize_srpt_timeline(timeline, release_dates, fig, ax, show_release_lines=True):
          """Create a linear visualization of the SRPT schedule"""
          # Define colors for jobs - more distinct colors
          colors = ['red', 'orange', 'blue', 'green', 'gray', 'purple', 'brown', 'pink', 'cyan', 'magenta']

          # Get max time for setting axis limits
          max_time = max(segment['end'] for segment in timeline)

          # Plot timeline segments
          for segment in timeline:
              start = segment['start']
              end = segment['end']
              job_id = segment['job_id']  # This could be a float in some cases

              # CHANGE THIS LINE - Explicitly convert job_id to int:
              color_idx = (int(job_id) - 1) % len(colors)
              color = colors[color_idx]

              # Create rectangle for this job segment
              rect = patches.Rectangle(
                  (start, 0),
                  end - start,
                  1,
                  linewidth=1,
                  edgecolor='black',
                  facecolor=color
              )
              ax.add_patch(rect)

              # Add job ID text inside the rectangle
              ax.text(
                start + (end - start)/2,
                0.5,
                f"J{int(job_id)}",
                color='white',
                fontweight='bold',
                ha='center',
                va='center',
                fontsize=10
            )

          # Identify preemption points and create flags
          job_segments = {}
          for segment in timeline:
              job_id = segment['job_id']
              if job_id not in job_segments:
                  job_segments[job_id] = []
              job_segments[job_id].append((segment['start'], segment['end']))

          # Look for non-contiguous segments for each job
          for job_id, segments in job_segments.items():
              # Sort segments by start time
              segments.sort()

              # Check for preemptions (non-contiguous segments)
              for i in range(len(segments) - 1):
                  if segments[i][1] != segments[i+1][0]:  # If end of segment != start of next segment
                      preemption_time = segments[i][1]

                      # Draw flag pole
                      ax.plot([preemption_time, preemption_time], [1, 1.5], 'k-', linewidth=1.5)

                      # Draw flag
                      flag = patches.Rectangle(
                          (preemption_time, 1.5),
                          1.5,
                          0.5,
                          linewidth=1,
                          edgecolor='black',
                          facecolor='darkgreen'
                      )
                      ax.add_patch(flag)

                      # Add job ID to flag
                      ax.text(
                          preemption_time + 0.75,
                          1.75,
                          f"J{int(job_id)}",
                          color='white',
                          fontweight='bold',
                          ha='center',
                          va='center',
                          fontsize=9
                      )

          # Add release time markers (red dashed lines) - only if toggle is enabled
          if show_release_lines:
              for t in sorted(set(release_dates)):
                  ax.axvline(x=t, color='red', linestyle='--', linewidth=1)

          # Set up the axis
          ax.set_xlim(-1, max_time + 3)
          ax.set_ylim(-0.2, 2.5)

          # Add time markers on x-axis
          time_range = np.arange(0, max_time + 4, 2)
          ax.set_xticks(time_range)
          ax.set_yticks([])

          # Add grid
          ax.grid(True, axis='x', linestyle='-', alpha=0.3)

          # Remove y-axis
          ax.spines['left'].set_visible(False)
          ax.spines['right'].set_visible(False)
          ax.spines['top'].set_visible(False)

          # Add title
          ax.set_title('SRPT (Shortest Remaining Processing Time) Schedule', fontsize=14)

          plt.tight_layout()

          return fig, ax

        # Run the algorithm
        SRPT(df)

    # 10. Branch and Bound Algorithm
    elif algorithm == "Branch and Bound (Minimize Maximum Lateness)":
        # Create placeholders for the algorithm results and visualization
        result_container = st.container()
        figure_container = st.container()
        download_container = st.container()

        def branch_and_bound_lmax(jobs_df):
            """
            Branch and Bound algorithm for 1|rj|Lmax using preemptive EDD lower bound
            with tracking of explored nodes
            """
            # Create a copy of the dataframe
            df = jobs_df.copy().reset_index(drop=True)

            # Create a node class for the branch and bound tree
            class Node:
                def __init__(self, level, sequence, completion_time, lmax, parent=None):
                    self.id = None  # Will be assigned during tree building
                    self.level = level
                    self.sequence = sequence.copy()
                    self.completion_time = completion_time
                    self.lmax = lmax  # Max lateness of the scheduled jobs
                    self.parent = parent
                    self.children = []
                    self.pruned = False
                    self.is_best = False
                    self.bound = float('inf')  # Store the bound separately
                    self.explored = False  # Track if the node was explored during search

                def __lt__(self, other):
                    return self.bound < other.bound

            # Check dominance rule: rj < min_{l∈J} (max(t,rl) + pl)
            def is_dominated(job_id, unscheduled, current_time):
                job = df[df['Job_ID'] == job_id].iloc[0]
                r_j = job['Release_Date']

                for other_id in unscheduled:
                    if other_id == job_id:
                        continue

                    other_job = df[df['Job_ID'] == other_id].iloc[0]
                    r_l = other_job['Release_Date']
                    p_l = other_job['Processing_Time']

                    # Check dominance condition
                    if r_j >= max(current_time, r_l) + p_l:
                        return True

                return False

            # Calculate preemptive EDD lower bound
            def preemptive_edd_bound(scheduled, current_time):
                # Calculate max lateness of scheduled jobs
                scheduled_lateness = 0
                if scheduled:
                    t = 0
                    for job_id in scheduled:
                        job = df[df['Job_ID'] == job_id].iloc[0]
                        t = max(t, job['Release_Date']) + job['Processing_Time']
                        lateness = t - job['Due_Date']
                        scheduled_lateness = max(scheduled_lateness, lateness)

                # Get unscheduled jobs
                unscheduled = [j for j in df['Job_ID'] if j not in scheduled]
                if not unscheduled:
                    return scheduled_lateness

                # Create job data for unscheduled jobs
                jobs = []
                for job_id in unscheduled:
                    job = df[df['Job_ID'] == job_id].iloc[0]
                    jobs.append({
                        'id': job_id,
                        'release': job['Release_Date'],
                        'processing': job['Processing_Time'],
                        'due': job['Due_Date'],
                        'remaining': job['Processing_Time']
                    })

                # Run preemptive EDD
                t = current_time
                lateness = scheduled_lateness

                while any(job['remaining'] > 0 for job in jobs):
                    # Get available jobs
                    available = [j for j in jobs if j['release'] <= t and j['remaining'] > 0]

                    if not available:
                        # Jump to next release time
                        next_time = min(j['release'] for j in jobs if j['remaining'] > 0)
                        t = next_time
                        continue

                    # Sort by EDD
                    available.sort(key=lambda x: x['due'])

                    # Process the job with earliest due date
                    current_job = available[0]

                    # Find the next event (release or completion)
                    next_releases = [j['release'] for j in jobs if j['release'] > t and j['remaining'] > 0]

                    if next_releases:
                        next_event = min(next_releases)
                        process_time = min(current_job['remaining'], next_event - t)
                        current_job['remaining'] -= process_time
                        t += process_time
                    else:
                        # No more releases, process to completion
                        t += current_job['remaining']
                        current_job['remaining'] = 0

                    # Update lateness when a job completes
                    if current_job['remaining'] == 0:
                        job_lateness = t - current_job['due']
                        lateness = max(lateness, job_lateness)

                return lateness

            # Create root node
            root = Node(level=0, sequence=[], completion_time=0, lmax=0)
            root.id = 0
            root.bound = preemptive_edd_bound([], 0)
            root.explored = True  # Root is always explored

            # Best solution found so far
            best_solution = None
            best_lmax = float('inf')

            # Use priority queue for best-first search
            queue = [root]
            heapq.heapify(queue)

            # Keep track of all nodes for visualization
            all_nodes = {0: root}
            next_id = 1

            # Start branch and bound
            with st.spinner("Running Branch and Bound algorithm..."):
                while queue:
                    current = heapq.heappop(queue)

                    # Mark this node as explored (popped from queue)
                    current.explored = True

                    # Skip if already pruned
                    if current.pruned:
                        continue

                    # If we've found a complete solution
                    if current.level == len(df):
                        if current.lmax < best_lmax:
                            if best_solution:
                                best_solution.is_best = False
                            best_lmax = current.lmax
                            best_solution = current
                            current.is_best = True

                            # Important: Retrospectively prune nodes with bounds >= best_lmax
                            for _, node in all_nodes.items():
                                if not node.is_best and node.bound >= best_lmax:
                                    node.pruned = True
                        continue

                    # Get unscheduled jobs
                    unscheduled = [j for j in df['Job_ID'] if j not in current.sequence]

                    # Generate children according to dominance rule
                    for job_id in unscheduled:
                        if is_dominated(job_id, unscheduled, current.completion_time):
                            continue

                        # Get job info
                        job = df[df['Job_ID'] == job_id].iloc[0]

                        # Calculate completion time
                        start_time = max(current.completion_time, job['Release_Date'])
                        completion_time = start_time + job['Processing_Time']

                        # Calculate maximum lateness of scheduled jobs
                        lateness = completion_time - job['Due_Date']
                        new_lmax = max(current.lmax, lateness)

                        # Create new sequence
                        new_sequence = current.sequence + [job_id]

                        # Create child node
                        child = Node(
                            level=current.level + 1,
                            sequence=new_sequence,
                            completion_time=completion_time,
                            lmax=new_lmax,
                            parent=current
                        )
                        child.id = next_id
                        next_id += 1

                        # Add child to parent
                        current.children.append(child)

                        # Store for visualization
                        all_nodes[child.id] = child

                        # Calculate lower bound using preemptive EDD
                        bound = preemptive_edd_bound(new_sequence, completion_time)
                        child.bound = max(new_lmax, bound)  # Store the bound

                        # Check if we can prune immediately
                        if child.bound >= best_lmax:
                            child.pruned = True
                            continue

                        # Add to queue
                        heapq.heappush(queue, child)

            # After finding the best solution, mark all nodes in the optimal path
            if best_solution:
                # Trace back from the optimal solution to the root
                current = best_solution
                while current:
                    current.is_best = True
                    current = current.parent

            # Create a tree for visualization
            G = nx.DiGraph()

            # Add nodes
            for node_id, node in all_nodes.items():
                # Generate node label
                if node.level == 0:
                    label = "∅"
                else:
                    label = str(node.sequence[-1])

                # Add node with attributes
                G.add_node(
                    node_id,
                    level=node.level,
                    sequence=node.sequence,
                    lmax=node.bound,  # Use bound as the displayed value
                    pruned=node.pruned,
                    is_best=node.is_best,
                    explored=node.explored,  # Include the explored flag
                    label=label
                )

            # Add edges
            for node_id, node in all_nodes.items():
                for child in node.children:
                    G.add_edge(node_id, child.id)

            # Return results
            if best_solution:
                return best_solution.sequence, best_lmax, G, all_nodes
            else:
                return None, float('inf'), G, all_nodes

        def calculate_optimal_schedule(optimal_sequence, jobs_df):
            """Calculate the actual schedule for the optimal sequence"""
            df = jobs_df.copy()
            current_time = 0
            job_schedule = []

            for job_id in optimal_sequence:
                job = df[df['Job_ID'] == job_id].iloc[0]

                # Calculate start time (considering release date)
                start_time = max(current_time, job['Release_Date'])

                # Calculate completion time
                completion_time = start_time + job['Processing_Time']

                # Calculate lateness
                lateness = completion_time - job['Due_Date']

                # Calculate tardiness
                tardiness = max(0, lateness)

                # Add to schedule
                job_schedule.append({
                    'Job_ID': job_id,
                    'Release_Date': job['Release_Date'],
                    'Processing_Time': job['Processing_Time'],
                    'Due_Date': job['Due_Date'],
                    'Start_Time': start_time,
                    'Completion_Time': completion_time,
                    'Lateness': lateness,
                    'Tardiness': tardiness
                })

                # Update current time
                current_time = completion_time

            return pd.DataFrame(job_schedule)

        def get_node_metrics(all_nodes):
            """Extract node metrics for analysis"""
            node_data = []

            for node_id, node in all_nodes.items():
                node_type = "Optimal Path" if node.is_best else "Active Path" if node.children else "Terminal"

                node_data.append({
                    'Node_ID': node_id,
                    'Level': node.level,
                    'Sequence': str(node.sequence),
                    'Bound': node.bound,
                    'Pruned': "Yes" if node.pruned else "No",
                    'Node_Type': node_type,
                    'Explored': "Yes" if node.explored else "No"
                })

            return pd.DataFrame(node_data)

        def visualize_branch_bound_tree(G, fig_size=(15, 10)):
            """
            Improved Branch and Bound tree visualization showing the optimal path
            """
            fig, ax = plt.subplots(figsize=fig_size)

            # Get node levels
            levels = nx.get_node_attributes(G, 'level')
            max_level = max(levels.values())

            # Group nodes by level
            nodes_by_level = {}
            for node, level in levels.items():
                if level not in nodes_by_level:
                    nodes_by_level[level] = []
                nodes_by_level[level].append(node)

            # Calculate positions with improved spacing
            pos = {}
            h_space = 3.0  # Horizontal spacing factor
            v_space = 2.0  # Vertical spacing factor

            for level in range(max_level + 1):
                if level in nodes_by_level:
                    nodes = sorted(nodes_by_level[level])
                    width = len(nodes)

                    # Apply spacing
                    for i, node in enumerate(nodes):
                        x = (i - width / 2 + 0.5) * h_space
                        y = -level * v_space
                        pos[node] = (x, y)

            # Identify different types of nodes - PATH-BASED CATEGORIZATION
            best_nodes = []           # Part of the optimal solution path
            path_nodes = []           # Nodes that led to further exploration but aren't optimal
            terminal_nodes = []       # Nodes that didn't lead anywhere (leaf nodes or pruned)

            for node in G.nodes():
                is_best = nx.get_node_attributes(G, 'is_best').get(node, False)

                if is_best:
                    best_nodes.append(node)
                else:
                    # Check if this node has children (successors)
                    successors = list(G.successors(node))

                    if successors:  # If node has children, it's part of the exploration path
                        path_nodes.append(node)
                    else:  # If no children, it's a terminal node
                        terminal_nodes.append(node)

            # Draw the graph in order

            # 1. Draw edges
            nx.draw_networkx_edges(G, pos, width=1.0, alpha=0.5)

            # 2. Draw terminal nodes
            if terminal_nodes:
                nx.draw_networkx_nodes(G, pos, nodelist=terminal_nodes,
                                      node_color='lightgray',
                                      node_size=900,
                                      alpha=0.9)

            # 3. Draw path nodes
            if path_nodes:
                nx.draw_networkx_nodes(G, pos, nodelist=path_nodes,
                                      node_color='skyblue',
                                      node_size=1000,
                                      alpha=1.0)

            # 4. Draw best solution nodes (ENTIRE optimal path)
            if best_nodes:
                nx.draw_networkx_nodes(G, pos, nodelist=best_nodes,
                                      node_color='lightgreen',
                                      node_size=1000,
                                      alpha=1.0)

            # Get attributes for labels
            labels = nx.get_node_attributes(G, 'label')
            sequences = nx.get_node_attributes(G, 'sequence')
            lmaxes = nx.get_node_attributes(G, 'lmax')

            # Create custom labels with sequence and bound
            node_labels = {}
            for node in G.nodes():
                sequence = sequences[node]
                bound = lmaxes[node]
                job_label = labels[node]

                if len(sequence) == 0:
                    node_labels[node] = f"∅\nBound={bound}"
                else:
                    node_labels[node] = f"Job {job_label}\nBound={bound}"

            # 5. Draw labels
            nx.draw_networkx_labels(G, pos, labels=node_labels, font_size=10)

            # Add a legend with matching colors
            legend_elements = [
                Line2D([0], [0], marker='o', color='w', markerfacecolor='skyblue', markersize=15,
                      label='Active Path Node'),
                Line2D([0], [0], marker='o', color='w', markerfacecolor='lightgray', markersize=15,
                      label='Terminal/Pruned'),
                Line2D([0], [0], marker='o', color='w', markerfacecolor='lightgreen', markersize=15,
                      label='Optimal Path')
            ]
            ax.legend(handles=legend_elements, loc='upper right')

            # Add a title
            ax.set_title("Branch and Bound Tree Visualization", fontsize=14)
            ax.axis('off')

            plt.tight_layout()
            return fig, ax

        def get_branch_bound_excel_download_link(schedule_df, summary_df, node_metrics_df, fig, algorithm_name):
            """Creates a downloadable Excel file for Branch and Bound results"""
            # Create a BytesIO object to store the Excel file
            output = io.BytesIO()

            # Create Excel writer
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Add each dataframe as a separate sheet
                schedule_df.to_excel(writer, sheet_name='Optimal Schedule', index=False)
                summary_df.to_excel(writer, sheet_name='Algorithm Summary', index=False)
                node_metrics_df.to_excel(writer, sheet_name='Search Tree Nodes', index=False)

                # Add an empty sheet for the tree visualization
                workbook = writer.book
                tree_sheet = workbook.create_sheet(title='B&B Tree Visualization')

                # Add a title to the tree sheet
                tree_sheet['A1'] = f"{algorithm_name} Branch and Bound Tree"

                # Apply styling to the title
                tree_sheet['A1'].font = Font(size=14, bold=True)
                tree_sheet['A1'].alignment = Alignment(horizontal='center')

                # Merge cells for the title
                tree_sheet.merge_cells('A1:G1')

                # Save the figure as a PNG in memory
                img_buf = io.BytesIO()
                fig.savefig(img_buf, format='png', dpi=150, bbox_inches='tight')
                img_buf.seek(0)

                # Create a PIL Image object from the buffer
                pil_img = PIL.Image.open(img_buf)

                # Create an openpyxl image
                xl_img = XLImage(img_buf)

                # Resize the image if needed
                scale_factor = 0.8
                xl_img.width = int(pil_img.width * scale_factor)
                xl_img.height = int(pil_img.height * scale_factor)

                # Add the image to the sheet
                tree_sheet.add_image(xl_img, 'A3')

                # Auto-adjust columns width for all sheets
                for sheet in workbook.sheetnames:
                    ws = workbook[sheet]
                    for column in ws.columns:
                        column_letter = get_column_letter(column[0].column)
                        if column_letter == 'A':
                            ws.column_dimensions[column_letter].width = 20
                        else:
                            ws.column_dimensions[column_letter].width = 15

            # Set buffer position to start
            output.seek(0)

            # Convert to base64 for download link
            b64 = base64.b64encode(output.getvalue()).decode()
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{algorithm_name}_results.xlsx">📊 Download All Results (Excel)</a>'

            return href

        # Run the Branch and Bound algorithm
        start_time = time.time()
        optimal_sequence, optimal_lmax, tree, all_nodes = branch_and_bound_lmax(df)
        end_time = time.time()
        computation_time = end_time - start_time

        # Calculate the optimal schedule based on the sequence
        optimal_schedule_df = calculate_optimal_schedule(optimal_sequence, df)

        # Extract node metrics for analysis
        node_metrics_df = get_node_metrics(all_nodes)

        # Count node types for statistics
        path_nodes = 0
        terminal_nodes = 0
        best_nodes = 0

        for node_id, node in all_nodes.items():
            if node.is_best:
                best_nodes += 1
            elif node.children:
                path_nodes += 1
            else:
                terminal_nodes += 1

        # Create summary dataframe
        summary_data = {
            'Metric': [
                'Optimal Sequence',
                'Minimum Maximum Lateness',
                'Computation Time (seconds)',
                'Active Path Nodes',
                'Terminal Nodes',
                'Optimal Path Nodes',
                'Total Nodes Explored'
            ],
            'Value': [
                str(optimal_sequence),
                optimal_lmax,
                f"{computation_time:.3f}",
                path_nodes,
                terminal_nodes,
                best_nodes,
                len(all_nodes)
            ]
        }
        summary_df = pd.DataFrame(summary_data)

        # Display results
        with result_container:
            st.subheader("Branch and Bound Results")
            st.write(f"Optimal sequence: {optimal_sequence}")
            st.write(f"Minimum maximum lateness: {optimal_lmax}")
            st.write(f"Computation time: {computation_time:.3f} seconds")
            st.write(f"Active path nodes: {path_nodes}")
            st.write(f"Terminal nodes: {terminal_nodes}")
            st.write(f"Optimal path nodes: {best_nodes}")

            # Show the optimal schedule
            st.subheader("Optimal Schedule:")
            st.dataframe(optimal_schedule_df)

        # Display visualizations
        with figure_container:
            # Visualize the branch and bound tree
            st.subheader("Branch and Bound Tree")
            fig1, ax1 = visualize_branch_bound_tree(tree)
            st.pyplot(fig1)

        # Add download options
        with download_container:
            st.subheader("Download Results")

            # Excel download with all data
            st.markdown(
                get_branch_bound_excel_download_link(
                    optimal_schedule_df,
                    summary_df,
                    node_metrics_df,
                    fig1,
                    "Branch_and_Bound"
                ),
                unsafe_allow_html=True
            )
